"""Parser for Socialstyrelsen metadata Excel deliveries.

Each Socialstyrelsen register is published as a standalone .xlsx workbook
with a consistent-but-not-uniform set of sheets. This module reads one
workbook and returns a `SosRegister` — a structured, DB-schema-independent
representation suitable for downstream DB ingestion or docs generation.

Known shape (derived from the 13 registers currently distributed):

    Generell information        — template & dataset version, contact
    Metadata-Datamängd (DCAT-AP) — register-level DCAT-AP metadata
    Deldatamängder och datavyer — subset/view descriptions (optional)
    Metadata - Variabelnivå     — variable rows (16 standard columns)
    Kodlista_*                  — per-variable value sets (optional)
    Kvalitet_*                  — free-form quality notes (LMED only)

Sheet names vary in case, whitespace, and punctuation; `_find_sheet`
matches on normalised tokens. Workbook files beginning with `~$` are
Microsoft Office lock files and are rejected up front.
"""

from __future__ import annotations

import calendar
import re
import zipfile
from collections import Counter
from dataclasses import dataclass, field, replace
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING, Any

if TYPE_CHECKING:
    from collections.abc import Iterable, Iterator
    from typing import Protocol

    from reg_meta_build.sources import IRObject

# ---------------------------------------------------------------------------
# Output types
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class SosDcatAp:
    """Register-level DCAT-AP metadata. All fields optional — older
    template versions or partial deliveries may omit any of them.

    `extras` holds rows whose Swedish attribute name we don't map to a
    known field, so the full sheet content survives parsing.
    """

    title_sv: str | None = None
    title_en: str | None = None
    description_sv: str | None = None
    description_en: str | None = None
    temporal_coverage_sv: str | None = None
    temporal_coverage_en: str | None = None
    geographic_coverage_sv: str | None = None
    geographic_coverage_en: str | None = None
    population_sv: str | None = None
    population_en: str | None = None
    update_frequency_sv: str | None = None
    update_frequency_en: str | None = None
    publisher_sv: str | None = None
    publisher_en: str | None = None
    contact_sv: str | None = None
    contact_en: str | None = None
    documentation_url_sv: str | None = None
    documentation_url_en: str | None = None
    landing_page_sv: str | None = None
    landing_page_en: str | None = None
    access_url_sv: str | None = None
    access_url_en: str | None = None
    access_rights_sv: str | None = None
    access_rights_en: str | None = None
    legislation_sv: str | None = None
    legislation_en: str | None = None
    extras: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class SosDeldatamangd:
    """One subset/view within a register. From `Deldatamängder…` sheet.

    Some registers (LSS, BU, SOL) lack this sheet entirely; the caller is
    expected to synthesise a single implicit deldatamängd for those.
    """

    name: str
    label: str | None
    description: str | None
    data_from: int | None
    data_to: int | None
    update_frequency: str | None
    aggregation_level: str | None


@dataclass(frozen=True)
class SosVariable:
    """One variable occurrence in a register (row in Metadata - Variabelnivå).

    Identity is `(deldatamangd, name)`. The same variable name can appear
    under multiple deldatamängder within the same register, and across
    registers — uniqueness is not guaranteed even within a single file.
    """

    deldatamangd: str | None
    name: str
    label: str | None
    description: str | None
    object_type: str | None
    value_set_text: str | None  # raw `Värdemängd` free-text
    external_classification: str | None  # raw `Länk kodverk`
    data_type: str | None
    is_join_variable: str | None
    join_description: str | None
    presentation_order: int | None
    data_from: int | None
    data_to: int | None
    quality_note: str | None
    origin: str | None
    source_detail: str | None


@dataclass(frozen=True)
class SosKodlistaRow:
    tidsperiod: str | None
    kod: str
    beskrivning: str | None
    variable_name: str | None = (
        None  # set only when sheet has a per-row Variabelnamn column
    )


@dataclass(frozen=True)
class SosKodlista:
    """Value set from a `Kodlista_*` sheet. Mapping to a variable is by
    sheet-name suffix (e.g. `Kodlista_DIAGNOS` → variable `DIAGNOS`).
    The caller is responsible for resolution — not guaranteed 1:1.

    `rows` holds structured (Tidsperiod, Kod, Beskrivning) entries. Sheets
    that don't match the standard header shape (recoding tables, hospital
    directories, ICD mapping tables etc.) parse with empty `rows` — the
    raw content is preserved in `raw_rows` for downstream custom handling.
    """

    sheet_name: str
    variable_hint: str  # suffix after `Kodlista_`
    codeset_name: str | None  # from "Kodverk" row, if present
    variable_header: str | None  # from "Variabelnamn" row, if present
    background: str | None  # from "Bakgrund" row, if present
    rows: tuple[SosKodlistaRow, ...]
    raw_rows: tuple[tuple[Any, ...], ...] = ()


@dataclass(frozen=True)
class SosQualitySheet:
    """A `Kvalitet_*` sheet captured verbatim. LMED uses these for
    register-level quality narrative. Rows are kept as raw tuples; no
    further structure is assumed."""

    sheet_name: str
    rows: tuple[tuple[Any, ...], ...]


@dataclass(frozen=True)
class SosRegister:
    source_file: Path
    dataset_name: str | None
    dataset_version: str | None
    dataset_date: date | None
    template_version: str | None
    template_date: date | None
    contact_email: str | None
    dcat_ap: SosDcatAp
    deldatamangder: tuple[SosDeldatamangd, ...]
    variables: tuple[SosVariable, ...]
    kodlistor: tuple[SosKodlista, ...]
    quality_sheets: tuple[SosQualitySheet, ...]
    warnings: tuple[str, ...]


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------


class SosParseError(Exception):
    """Raised when the workbook cannot be read or is missing required sheets."""


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


def parse_register_file(path: Path | str) -> SosRegister:
    """Read one Socialstyrelsen register workbook and return structured
    metadata. Raises `SosParseError` on unreadable / unrecognised files."""

    import openpyxl
    import openpyxl.utils.exceptions

    p = Path(path)
    if p.name.startswith("~$"):
        raise SosParseError(f"{p.name} is an Office lock file; skip")
    if not p.is_file():
        raise SosParseError(f"{p} is not a regular file (missing or a directory)")
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except zipfile.BadZipFile as exc:
        raise SosParseError(f"{p.name} is not a valid .xlsx file") from exc
    except openpyxl.utils.exceptions.InvalidFileException as exc:
        # `.xls`, `.xlsb`, and other formats openpyxl doesn't support.
        raise SosParseError(
            f"{p.name}: openpyxl does not support this file format "
            "(only .xlsx/.xlsm/.xltx/.xltm)"
        ) from exc
    except (OSError, ValueError, KeyError) as exc:
        # openpyxl can raise these on partially corrupt files (truncated XML,
        # missing relationships, unexpected schema). Wrap so callers see a
        # uniform error type.
        raise SosParseError(
            f"{p.name} could not be read as a valid .xlsx file: {exc}"
        ) from exc

    try:
        warnings: list[str] = []
        norm_sheets = {_normalise(n): n for n in wb.sheetnames}

        generell = _find_sheet(norm_sheets, ["generell", "information"])
        dcat = _find_sheet(norm_sheets, ["datamängd", "dcat"]) or _find_sheet(
            norm_sheets, ["metadata", "datamängd"]
        )
        deldat = (
            _find_sheet(norm_sheets, ["deldatamängder", "datavyer"])
            or _find_sheet(norm_sheets, ["metadata", "deldatamängder"])
            or _find_sheet(norm_sheets, ["deldatamängder"])
        )
        varsheet = _find_sheet(
            norm_sheets, ["metadata", "variabelnivå"]
        ) or _find_sheet(norm_sheets, ["metadata", "variabler"])

        if varsheet is None:
            raise SosParseError(f"{p.name}: no variable-level sheet found")

        gen = _parse_generell(wb[generell]) if generell else {}
        dcat_ap = _parse_dcat_ap(wb[dcat]) if dcat else SosDcatAp()
        deldatamangder = tuple(_parse_deldatamangder(wb[deldat])) if deldat else ()
        variables = tuple(_parse_variables(wb[varsheet]))

        kodlistor: list[SosKodlista] = []
        quality_sheets: list[SosQualitySheet] = []
        for sheet_name in wb.sheetnames:
            low = sheet_name.lower()
            if low.startswith("kodlista"):
                try:
                    kod, kod_warnings = _parse_kodlista(wb[sheet_name])
                    kodlistor.append(kod)
                    warnings.extend(kod_warnings)
                except Exception as exc:  # noqa: BLE001 — best-effort parse boundary: any sheet failure downgrades to a raw-hint warning
                    warnings.append(f"kodlista {sheet_name!r}: {exc}")
                    # Preserve kodlista-wins (#401): a sheet that FAILS to parse is
                    # still a kodlista sheet — record its hint as a raw/unparseable
                    # placeholder so the Värdemängd fallback won't fabricate codes
                    # for its variable. Flows through the SAME raw_rows skip path as
                    # a genuinely raw sheet (excluded from value-set construction in
                    # _emit_register, counted in raw_kodlista_hints). Suffix derived
                    # exactly as in _parse_kodlista. (0 corpus occurrences today.)
                    hint = (
                        sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
                    )
                    hint = hint.split("!", 1)[0].strip()
                    kodlistor.append(
                        SosKodlista(
                            sheet_name=sheet_name,
                            variable_hint=hint,
                            codeset_name=None,
                            variable_header=None,
                            background=None,
                            rows=(),
                            raw_rows=(("<unparseable: parse error>",),),
                        )
                    )
            elif low.startswith("kvalitet"):
                quality_sheets.append(_parse_quality_sheet(wb[sheet_name]))

        if generell is None:
            warnings.append("missing Generell information sheet")
        if dcat is None:
            warnings.append("missing DCAT-AP sheet")
        if deldat is None:
            warnings.append("missing Deldatamängder sheet (implicit single subset)")

        return SosRegister(
            source_file=p,
            dataset_name=gen.get("dataset_name"),
            dataset_version=gen.get("dataset_version"),
            dataset_date=gen.get("dataset_date"),
            template_version=gen.get("template_version"),
            template_date=gen.get("template_date"),
            contact_email=gen.get("contact_email"),
            dcat_ap=dcat_ap,
            deldatamangder=deldatamangder,
            variables=variables,
            kodlistor=tuple(kodlistor),
            quality_sheets=tuple(quality_sheets),
            warnings=tuple(warnings),
        )
    finally:
        wb.close()


def parse_directory(directory: Path | str) -> list[SosRegister]:
    """Parse every `.xlsx` file in a directory, skipping Office lock files.
    Halts on the first parse failure (raises `SosParseError`); call per file
    if you need to collect errors instead."""

    d = Path(directory)
    out: list[SosRegister] = []
    for f in sorted(d.iterdir()):
        if not f.is_file():
            continue
        # Case-insensitive: some deliveries arrive as `.XLSX` on case-sensitive
        # filesystems, and a strict `*.xlsx` glob would skip them silently.
        if f.suffix.lower() != ".xlsx":
            continue
        if f.name.startswith("~$"):
            continue
        out.append(parse_register_file(f))
    return out


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------


def _normalise(s: str) -> str:
    return re.sub(r"[\s_\-()]+", "", s).lower()


def _find_sheet(norm_sheets: dict[str, str], tokens: list[str]) -> str | None:
    """Return the first original sheet name whose normalised form contains
    every token in `tokens` (also normalised). Caller is expected to build
    `norm_sheets` once via `{_normalise(n): n for n in wb.sheetnames}` so
    repeated lookups don't re-normalise."""
    wanted = [_normalise(t) for t in tokens]
    for norm, original in norm_sheets.items():
        if all(t in norm for t in wanted):
            return original
    return None


def _row_iter(ws: Any, start: int = 1) -> Iterator[tuple[Any, ...]]:
    """Yield rows starting at `start`, stopping after a long empty tail.
    openpyxl's `max_row` is unreliable (phantom rows in some deliveries)."""
    empty_streak = 0
    empty_limit = 50
    for row in ws.iter_rows(min_row=start, values_only=True):
        if any(v is not None and str(v).strip() for v in row):
            empty_streak = 0
            yield row
        else:
            empty_streak += 1
            if empty_streak >= empty_limit:
                break


def _cell_row_iter(ws: Any, start: int = 1) -> Iterable[tuple[Any, ...]]:
    """Like `_row_iter` but yields tuples of openpyxl cell objects, so
    callers can inspect formatting (e.g. number_format on code columns)."""
    empty_streak = 0
    empty_limit = 50
    for cells in ws.iter_rows(min_row=start, values_only=False):
        if any(c.value is not None and str(c.value).strip() for c in cells):
            empty_streak = 0
            yield tuple(cells)
        else:
            empty_streak += 1
            if empty_streak >= empty_limit:
                break


def _at(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _pick(row: tuple[Any, ...], col_map: dict[str, int], field_name: str) -> Any:
    """Look up `field_name` in `col_map` and return the row value at that
    index, or None if the column is absent or short. Convenience for
    header-mapped sheet parsers."""
    return _at(row, col_map.get(field_name))


_PURE_ZERO_FMT = re.compile(r"^0+$")


def _format_code(cell: Any) -> str | None:
    """Render a code-column cell to a string, preserving leading zeros from
    Excel display formatting. Excel may store '001' as the integer 1 with
    number_format '000'; without consulting the format we'd silently emit
    '1' and corrupt code identity for downstream joins."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):  # bool is a subclass of int — handle first
        return str(v)
    if isinstance(v, (int, float)):
        if isinstance(v, float):
            if not v.is_integer():
                return str(v)
            v = int(v)
        fmt = cell.number_format or ""
        if _PURE_ZERO_FMT.fullmatch(fmt):
            return str(v).zfill(len(fmt))
        return str(v)
    s = str(v).strip()
    return s or None


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v)


def _as_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else None
    try:
        return int(str(v).strip())
    except TypeError, ValueError:
        return None


def _as_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---------------------------------------------------------------------------
# Sheet-specific parsers
# ---------------------------------------------------------------------------


def _parse_generell(ws: Any) -> dict[str, Any]:
    """Scan `Generell information`. Layout is key–value pairs scattered
    across roughly 25 rows with section headings; we key on the label in
    column B (index 1) and read the value from column C (index 2)."""

    out: dict[str, Any] = {}
    rows = list(ws.iter_rows(values_only=True))
    section: str | None = None
    for row in rows:
        if len(row) < 2:
            continue
        label = _clean(row[1]) if len(row) > 1 else None
        value = _clean(row[2]) if len(row) > 2 else None
        raw_value = row[2] if len(row) > 2 else None

        if label and value is None:
            low = label.lower()
            # Some deliveries have "metadatat" (typo) instead of "metadata";
            # match on the distinguishing tail words instead of the exact phrase.
            if "metadatamallen" in low:
                section = "template"
            elif "datamängden" in low and "version" in low:
                section = "dataset"
            continue

        if not label or value is None:
            continue

        low = label.lower()
        if section == "template" and low.startswith("version"):
            out["template_version"] = value
        elif section == "template" and low.startswith("datum"):
            out["template_date"] = _as_date(raw_value)
        elif section == "dataset" and low.startswith("datamängd"):
            out["dataset_name"] = value
        elif section == "dataset" and low.startswith("version"):
            out["dataset_version"] = value
        elif section == "dataset" and low.startswith("datum"):
            out["dataset_date"] = _as_date(raw_value)
        elif "e-post" in low or low == "e-post:":
            out["contact_email"] = value
    return out


# DCAT-AP attribute (Swedish) → internal field stem. We store both SV and
# EN columns as separate `<stem>_sv` / `<stem>_en` fields.
_DCAT_MAP = {
    "titel": "title",
    "beskrivning": "description",
    "tidsperiod": "temporal_coverage",
    "namngivet geografiskt område": "geographic_coverage",
    "population": "population",
    "uppdateringsfrekvens": "update_frequency",
    "utgivare": "publisher",
    "kontaktuppgift": "contact",
    "dokumentation": "documentation_url",
    "ingångssida": "landing_page",
    "webbadress för åtkomst": "access_url",
    "åtkomsträttigheter": "access_rights",
    "tillämplig lagstiftning": "legislation",
}


def _parse_dcat_ap(ws: Any) -> SosDcatAp:
    fields: dict[str, str | None] = {}
    extras: dict[str, str] = {}
    first = True
    for row in _row_iter(ws):
        if first:
            first = False
            # First row is column headers (Attribut SoS-metadata | ... | Svenska | Engelska)
            continue
        if len(row) < 3:
            continue
        attr = _clean(row[0])
        sv = _clean(row[2]) if len(row) > 2 else None
        en = _clean(row[3]) if len(row) > 3 else None
        if not attr:
            continue
        stem = _DCAT_MAP.get(attr.lower())
        if stem is None:
            # Capture unrecognised rows for inspection; value preference SV > EN
            value = sv or en or ""
            if value:
                extras[attr] = value
            continue
        if sv is not None:
            fields[f"{stem}_sv"] = sv
        if en is not None:
            fields[f"{stem}_en"] = en
    return SosDcatAp(**fields, extras=extras)


_DELDATAMANGD_HEADERS = {
    "deldatamängdsnamn": "name",
    "deldatamängdsetikett": "label",
    "deldatamängbeskrivning": "description",
    "deldatamängdsbeskrivning": "description",
    "data från": "data_from",
    "data till": "data_to",
    "uppdateringsfrekvens": "update_frequency",
    "aggregeringsnivå": "aggregation_level",
}


def _parse_deldatamangder(ws: Any) -> Iterable[SosDeldatamangd]:
    rows = _row_iter(ws)
    header = next(rows, None)
    if header is None:
        return
    col_map: dict[str, int] = {}
    for i, h in enumerate(header):
        cleaned = _clean(h) if h else None
        stem = _DELDATAMANGD_HEADERS.get(cleaned.lower() if cleaned else "")
        if stem:
            col_map[stem] = i

    if "name" not in col_map:
        return  # not a deldatamängd sheet shape; silently skip

    for row in rows:
        name = _clean(_pick(row, col_map, "name"))
        if not name:
            continue
        yield SosDeldatamangd(
            name=name,
            label=_clean(_pick(row, col_map, "label")),
            description=_clean(_pick(row, col_map, "description")),
            data_from=_as_int(_pick(row, col_map, "data_from")),
            data_to=_as_int(_pick(row, col_map, "data_to")),
            update_frequency=_clean(_pick(row, col_map, "update_frequency")),
            aggregation_level=_clean(_pick(row, col_map, "aggregation_level")),
        )


_VAR_HEADERS = {
    "deldatamängdsnamn": "deldatamangd",
    # BU splits deldatamängd into dataset + view; we keep the view name
    # ("Datavynamn") as the deldatamängd identity and drop the parent
    # ("Datamängdsnamn") since it duplicates the register-level name.
    "datavynamn": "deldatamangd",
    "variabelnamn": "name",
    "variabeletikett": "label",
    "variabelbeskrivning": "description",
    "objekttyp": "object_type",
    "värdemängd": "value_set_text",
    "länk kodverk": "external_classification",
    "datatyp": "data_type",
    "kopplingsvariabel": "is_join_variable",
    "kopplingsbeskrivning": "join_description",
    "presentationsordning": "presentation_order",
    "data från": "data_from",
    "data till": "data_to",
    "kvalitetsanmärkning": "quality_note",
    "ursprung": "origin",
    "specificera källa": "source_detail",
}


def _parse_variables(ws: Any) -> Iterable[SosVariable]:
    rows = _row_iter(ws)
    header = next(rows, None)
    if header is None:
        raise SosParseError(
            f"variable sheet {ws.title!r} is empty; cannot extract variables"
        )
    col_map: dict[str, int] = {}
    for i, h in enumerate(header):
        if not h:
            continue
        cleaned = _clean(h)
        if cleaned is None:
            continue
        stem = _VAR_HEADERS.get(cleaned.lower())
        if stem:
            col_map[stem] = i

    if "name" not in col_map:
        # Without a Variabelnamn column we silently return zero rows, hiding
        # an upstream rename or malformed delivery. Fail loudly instead.
        header_cols = ", ".join(repr(h) for h in header if h) or "(none)"
        raise SosParseError(
            f"variable sheet {ws.title!r} is missing a 'Variabelnamn' header; "
            f"found columns: {header_cols}"
        )

    for row in rows:
        name = _clean(_pick(row, col_map, "name"))
        if not name:
            continue
        yield SosVariable(
            deldatamangd=_clean(_pick(row, col_map, "deldatamangd")),
            name=name,
            label=_clean(_pick(row, col_map, "label")),
            description=_clean(_pick(row, col_map, "description")),
            object_type=_clean(_pick(row, col_map, "object_type")),
            value_set_text=_clean(_pick(row, col_map, "value_set_text")),
            external_classification=_clean(
                _pick(row, col_map, "external_classification")
            ),
            data_type=_clean(_pick(row, col_map, "data_type")),
            is_join_variable=_clean(_pick(row, col_map, "is_join_variable")),
            join_description=_clean(_pick(row, col_map, "join_description")),
            presentation_order=_as_int(_pick(row, col_map, "presentation_order")),
            data_from=_as_int(_pick(row, col_map, "data_from")),
            data_to=_as_int(_pick(row, col_map, "data_to")),
            quality_note=_clean(_pick(row, col_map, "quality_note")),
            origin=_clean(_pick(row, col_map, "origin")),
            source_detail=_clean(_pick(row, col_map, "source_detail")),
        )


def _parse_kodlista(ws: Any) -> tuple[SosKodlista, list[str]]:
    """A Kodlista sheet has a preamble (rows labelled Kodverk / Variabelnamn
    / Bakgrund) then a header row with (Tidsperiod, Kod, Beskrivning) and
    data rows beneath. Some sheets omit the preamble."""

    codeset_name: str | None = None
    variable_header: str | None = None
    background: str | None = None
    data_rows: list[SosKodlistaRow] = []
    raw_rows: list[tuple[Any, ...]] = []
    warnings: list[str] = []

    sheet_name = ws.title
    suffix = sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
    suffix = suffix.split("!", 1)[0].strip()

    col_tp: int | None = None
    col_kod: int | None = None
    col_desc: int | None = None
    col_var: int | None = None
    # forward-fill: some sheets put the period once on a header row above
    # rows that leave it blank
    last_tidsperiod: str | None = None
    # iterate cells (not just values) so the kod column can preserve leading
    # zeros from number_format
    all_cell_rows = list(_cell_row_iter(ws))
    for cells in all_cell_rows:
        row = tuple(c.value for c in cells)
        first = _clean(row[0]) if row else None
        second = _clean(row[1]) if len(row) > 1 else None

        if col_tp is None:
            # Detect the column-header row by looking for "Tidsperiod" + "Kod"
            # anywhere in the row. MFR-style sheets also carry a Variabelnamn
            # column (per-row variable) in addition to the Kodverk preamble
            # shape used by PAR et al.
            positions: dict[str, int] = {}
            for i, h in enumerate(row):
                hl = (_clean(h) or "").lower()
                if hl.startswith("tidsperiod"):
                    positions["tp"] = i
                elif hl == "kod":
                    positions["kod"] = i
                elif hl.startswith(("beskrivning", "betydelse")):
                    positions["desc"] = i
                elif hl == "variabelnamn":
                    positions["var"] = i
            if "tp" in positions and "kod" in positions:
                col_tp = positions["tp"]
                col_kod = positions["kod"]
                col_desc = positions.get("desc")
                col_var = positions.get("var")
                continue

            # Preamble rows (PAR-style): "Kodverk", "Variabelnamn", "Bakgrund"
            # appear in col 0 with the value in col 1.
            if first:
                key = first.lower()
                if key == "kodverk":
                    codeset_name = second
                    continue
                if key == "variabelnamn":
                    variable_header = second
                    continue
                if key == "bakgrund":
                    background = second
                    continue
            continue

        tp_val = _clean(_at(row, col_tp))
        kod_str = (
            _format_code(cells[col_kod])
            if col_kod is not None and col_kod < len(cells)
            else None
        )

        # Rows carrying only a Tidsperiod (no code) act as a section header
        # for the rows beneath them; remember and forward-fill.
        if tp_val and not kod_str:
            last_tidsperiod = tp_val
            continue
        if not kod_str:
            continue

        data_rows.append(
            SosKodlistaRow(
                tidsperiod=tp_val or last_tidsperiod,
                kod=kod_str,
                beskrivning=_clean(_at(row, col_desc)),
                variable_name=(
                    _clean(_at(row, col_var)) if col_var is not None else None
                ),
            )
        )

    if col_tp is None:
        warnings.append(
            f"kodlista {sheet_name!r}: no Tidsperiod/Kod header row found; "
            "structured rows skipped (raw content preserved)"
        )
        raw_rows = [tuple(c.value for c in cells) for cells in all_cell_rows]

    return (
        SosKodlista(
            sheet_name=sheet_name,
            variable_hint=suffix,
            codeset_name=codeset_name,
            variable_header=variable_header,
            background=background,
            rows=tuple(data_rows),
            raw_rows=tuple(raw_rows),
        ),
        warnings,
    )


def _parse_quality_sheet(ws: Any) -> SosQualitySheet:
    rows: list[tuple[Any, ...]] = []
    for row in _row_iter(ws):
        rows.append(tuple(row))
    return SosQualitySheet(sheet_name=ws.title, rows=tuple(rows))


# ---------------------------------------------------------------------------
# SOSAdapter (A4.3b) — Socialstyrelsen on the Model A materializer
# ---------------------------------------------------------------------------
#
# The adapter turns the parser's `SosRegister` trees into the provider-neutral
# IR stream the materializer (`reg_meta_build.db`) consumes. It is PURELY
# ADDITIVE: every `*_id` is hash-`mint()`ed into the `[2^62, 2^63)` band
# (disjoint from SCB's source-derived low band), and value_sets content-SHARE
# SCB's rows by member_hash (no minting). A `--providers=scb` build never
# instantiates this adapter, so the SCB-only DB stays byte-identical.
#
# R4 (verified against the live 13 workbooks at
# `reg_meta_build/input_data/Socialstyrelsen/`):
#   - <abbrev>: the parenthesized code in the workbook filename stem
#     ("Metadata ... (LMED)_webb.xlsx" -> "LMED"), lowercased to the mint token
#     "lmed". All 13 derive a unique uppercase code; lowercasing matches the
#     provider-slug convention and keeps the allow-list keys readable.
#   - SPLIT allow-list keys (register_abbrev, var.name) — the two known
#     same-name conflicts with a clean structured seam, both data_type-only
#     (no codelist), so they split on normalized data_type:
#       ("bu", "FOD_DATUMN")  Datum vs Heltal   (Insatser till barn och unga)
#       ("par", "ATC")        Sträng (text) vs Heltal (Patientregistret)
#     Every OTHER same-name conflict (BU AVSLDAT/BESLDAT/..., PAR ATCO/FODDAT/...)
#     is NOT allow-listed -> warn-merge (fail-soft).
#   - MFR entity-registry key (register_abbrev, SosKodlista.variable_hint):
#       ("mfr", "IVF_klinik")  (Medicinska födelseregistret) — 19 codes each
#     bound to ONE of 15 tidsperiods -> collapse to one state + per-code
#     valid_from/to. variable_hint preserves the sheet-suffix case ("IVF_klinik")
#     while the variable name is "IVF_KLINIK"; hint->variable match is
#     case-insensitive.

# A4.3b imports the IR contract + shared hashing/mint/IO infra from the build
# core. Safe against the db<->sources cycle: `db.py` imports the SOS adapter
# only function-locally (in `build_db`), never at module top.
from reg_meta_build.db import (
    _VALID_FROM_UNKNOWN,
    _file_sha256,
    _value_set_hash,
)
from reg_meta_build.id import mint
from reg_meta_build.ir import (
    IRDeliveryProvenance,
    IRRegister,
    IRValueCode,
    IRVariable,
    IRVariableAlias,
    IRVariableState,
    IRVariant,
    IRWarning,
)

# (register_abbrev, var.name) groups that auto-SPLIT into disjoint siblings.
# Fail-soft: only these known conflicts split; any other same-name conflict
# warn-merges (see SOSAdapter._resolve_variable_group). R4-verified above.
KNOWN_SPLIT_ALLOWLIST: frozenset[tuple[str, str]] = frozenset(
    {("bu", "FOD_DATUMN"), ("par", "ATC")}
)

# (register_abbrev, var.name) same-name MERGES that are intentional and
# type-lossless, so they should NOT emit `sos_unanticipated_same_name_conflict`.
# These keys still MERGE exactly as before — the guard only silences the warn.
# Data_type lives on `variable_state`, so a divergent data_type across members is
# preserved per state; the merge loses no information. #362.
#   - LOVA EXAMAR: `Examensår` (Heltal, A_LOVA & A_LOVA_HOSP) vs
#     `Utbildningsår (avslutningsår högsta utb.)` (Sträng (text), A_LOVA_EXAMEN).
#     Real-data verified against input_data/Socialstyrelsen/ (the LOVA workbook).
KNOWN_MERGE_ALLOWLIST: frozenset[tuple[str, str]] = frozenset({("lova", "EXAMAR")})

# Curated single-row variable-name corrections, keyed on the EXACT
# (register_abbrev, name, etikett) tuple — NOT a name pattern (the standing
# no-regex/name-pattern curation rule). Fixes an upstream typo where two
# DISTINCT variables ship under one name, disambiguated only by their etikett
# (`label`); the correction re-keys the mistyped row so the two no longer merge.
# #362.
#   - LOVA A_LOVA_PERSON ships INVARN1..INVARN9 (numeric immigration dates), but
#     the 9th event is mistyped as a SECOND INVARN8 (both `Heltal`, both
#     A_LOVA_PERSON). They share name AND normalized data_type, so the conflict
#     detector never fires and they merge SILENTLY, dropping invarn9. The two
#     rows differ ONLY in etikett: 'Invandringsdatum 8 numerisk' (the real
#     INVARN8) vs 'Invandringsdatum 9 numerisk' (the mistyped row → INVARN9),
#     so the etikett is the disambiguator. Real-data verified against the LOVA
#     workbook at input_data/Socialstyrelsen/.
VARIABLE_NAME_CORRECTIONS: dict[tuple[str, str, str], str] = {
    ("lova", "INVARN8", "Invandringsdatum 9 numerisk"): "INVARN9",
}

# (register_abbrev, SosKodlista.variable_hint) kodlistor that are ENTITY
# REGISTRIES (a stable directory whose entries each have their own active
# window), not value-set drift. Collapse to ONE state with per-code
# valid_from/to instead of one state per tidsperiod. R4-verified above.
ENTITY_REGISTRY_KODLISTOR: frozenset[tuple[str, str]] = frozenset(
    {("mfr", "IVF_klinik")}
)


def _is_styrtabell(d: SosDeldatamangd) -> bool:
    """A styrtabell (value-set decode table), not a research data subset.

    Requires BOTH structural signals (they agree across every current SOS
    delivery): the controlled-vocab Aggregeringsnivå == 'Ej relevant' AND the
    'Styrtabell …' Deldatamängdsetikett. Decode tables are excluded from
    variant/variable minting (#373) — their rows are data + point-in-time
    klartext, bound to the coded variable's value set, not minted as variables.
    """
    agg = (d.aggregation_level or "").strip().casefold() == "ej relevant"
    etikett = (d.label or "").strip().casefold().startswith("styrtabell")
    return agg and etikett


# Curated variable-sheet deldatamängd token -> Deldatamängder-sheet row name(s)
# (#211, the A4.4 curation the `sos_deldatamangd_unresolved` warning deferred).
# Four workbooks key their variable rows on a TECHNICAL extraction/view token
# that never appears as a Deldatamängder-sheet name; without this map their
# members warn-drop and the variants stay stateless. EXACT tokens only (the
# standing curation rule: no regex/name-pattern inference) keyed on
# (register_abbrev, variable-sheet token). A token can name SEVERAL variants
# (LMED's combined token) — the member emits a state/alias into EACH.
#
# This map is the ONLY bridge between the Variabelnivå sheet's technical tokens
# and the Deldatamängder sheet's display names (SOS tokens never equal the sheet
# names), so it is load-bearing for ALL listed tokens — including the styrtabell
# entries below, which the styrtabell exclusion (#373) reuses to resolve a
# variable row's deldatamängd token to its sheet name. styrtabell deldatamängder
# are detected via `_is_styrtabell` and EXCLUDED from BOTH variant and variable
# minting, so decode tables (kod->klartext) don't surface as research variables;
# their `A_LOVA_STYR_*` entries REMAIN here precisely because that exclusion uses
# them as the token->name bridge.
#
# Verified against the live workbooks at input_data/Socialstyrelsen/:
#   - LVM: tokens are the lowercase technical `label`s of the Deldatamängder
#     rows (lvm_ansok <-> 'LVM_ANSOK'), names are the long Swedish titles.
#   - LOVA: tokens are A_LOVA* extraction-table names. The workbook ships TWO
#     'LOVA' Deldatamängder rows (arbetsmarknadsstatus + the LISA-derived
#     ekonomi subset) that dedup to ONE minted variant — both A_LOVA and
#     A_LOVA_LISA land there. The A_LOVA_STYR_* entries name the styrtabell
#     decode tables; they stay mapped (the #373 exclusion needs them) but the
#     deldatamängder they point at are excluded from minting.
#   - DORS: the Covid-19 Hermes view's 26 variable rows use 'DORS-COV'.
#   - LMED: FDDD ships in BOTH variants; its token names them combined.
DELDATAMANGD_TOKEN_MAP: dict[tuple[str, str], tuple[str, ...]] = {
    ("lvm", "lvm_ansok"): (
        "Ansökningar om tvångsvård enligt lagen om vård av missbrukare "
        "i vissa fall, LVM",
    ),
    ("lvm", "lvm_omhtg"): (
        "Beslut om omedelbart omhändertagande för tvångsvård enligt lagen "
        "om vård av missbrukare i vissa fall, LVM",
    ),
    ("lvm", "lvm_utskr"): (
        "Tvångsvård enligt lagen om vård av missbrukare i vissa fall, LVM",
    ),
    ("lova", "A_LOVA"): ("LOVA",),
    ("lova", "A_LOVA_LISA"): ("LOVA",),
    ("lova", "A_LOVA_EXAMEN"): ("LOVA EXAMEN",),
    ("lova", "A_LOVA_HOSP"): ("LOVA HOSP",),
    ("lova", "A_LOVA_PERSON"): ("LOVA PERSON",),
    ("lova", "A_LOVA_STYR_AGARKAT"): ("LOVA AGARKAT",),
    ("lova", "A_LOVA_STYR_ARB_MARK_STATUS"): ("LOVA ARB_MARK_STAT",),
    ("lova", "A_LOVA_STYR_EXAMENSKODER"): ("LOVA EXAMENSKODER",),
    ("lova", "A_LOVA_STYR_HOSP_KODER"): ("LOVA HOSPKODER",),
    ("lova", "A_LOVA_STYR_LANDSKOD"): ("LOVA LANDSKOD",),
    ("lova", "A_LOVA_STYR_LEG_SPEC_EXAM_KOD"): ("LOVA LEG_SPEC_EXAM",),
    ("lova", "A_LOVA_STYR_REGION"): ("LOVA REGION",),
    ("lova", "A_LOVA_STYR_SEKTORKOD"): ("LOVA SEKTORKOD",),
    ("lova", "A_LOVA_STYR_SYSSSTAT"): ("LOVA SYSSSTAT",),
    ("lova", "A_LOVA_STYR_YRKSTALLN"): ("LOVA YRKSTALLN",),
    ("dors", "DORS-COV"): ("COV_DORS_HERMES",),
    ("lmed", "LMED VARA/LMED"): ("LMED", "LMED VARA"),
}

# Curated `Länk kodverk` (external_classification) → classification short_name.
#
# SOS has no `value_set_version_label`; the code system a variable uses is named
# only in the free-text `external_classification` field. This is a CURATED map of
# the OFFICIAL signals that unambiguously name a seeded classification (PR1) —
# NOT a broad regex/name-guesser. Each tuple is (substring-signal, short_name);
# a value matches when it CONTAINS the signal (case-insensitively), so trailing
# slashes / query strings / a bundled-PDF value that also carries the kva
# fragment all resolve. Order is irrelevant (signals are disjoint across systems).
#
# Everything NOT listed → None (unresolved): SCB LKF/SSYK/SUN/SNI URLs,
# skatteverket landskoder, TNM/Wiley, Op6-only PDFs, SOSNYK, sjukhuskoder,
# postnummer, all `Kodlista_*`/sheet-refs, prose, and typos. Real-data values
# verified against the 13 workbooks at input_data/Socialstyrelsen/.
_CLASSIFICATION_SIGNALS: tuple[tuple[str, str], ...] = (
    ("icd.who.int/browse10", "ICD-10-SE"),
    ("klassifikationer-och-koder/icd-10/", "ICD-10-SE"),
    ("icd.who.int/browse/releases", "ICD-10-SE"),
    ("fass.se/lif/atcregister", "ATC"),
    ("atcddd.fhi.no", "ATC"),
    ("klassifikationer-och-koder/kva/", "KVA"),
    ("klassifikation av vårdåtgärder (kvå) - socialstyrelsen", "KVA"),
    ("klassifikationer-och-koder/drg/", "DRG"),
)

# CAN (Cancerregistret) variable-name override. CAN's `external_classification`
# is a single globalassets PDF / Wiley TNM link that names NO seeded system, so
# the signal map yields None; the real code system is named by the VARIABLE name
# instead. Only ICD9 maps to a seeded classification (ICD-9-KS87); the historical
# tumour/morphology systems (ICD-7, ICD-O, SNOMED, TNM M/N/T, MORF) are not
# seeded → explicit None (documents the decision; the default is None anyway).
_CAN_VARNAME_CLASSIFICATION: dict[str, str | None] = {
    "ICD9": "ICD-9-KS87",
    "ICD7": None,
    "ICDO3": None,
    "ICDO10": None,
    "SNOMED3": None,
    "SNOMEDO10": None,
    "M": None,
    "N": None,
    "T": None,
    "MORF": None,
}


def _resolve_by_signal(external_classification: str | None) -> str | None:
    """Match an `external_classification` value against the curated signal map
    (substring, case-insensitive). Returns the classification short_name or None.
    This is the URL/string-signal half of the resolver — distinct from the CAN
    variable-name override."""
    if not external_classification:
        return None
    haystack = external_classification.lower()
    for signal, short_name in _CLASSIFICATION_SIGNALS:
        if signal in haystack:
            return short_name
    return None


def _resolve_classification(
    external_classification: str | None,
    abbrev: str,
    var_name: str,
) -> str | None:
    """Resolve a SOS variable to a seeded classification `short_name`, or None.

    Precedence: the `external_classification` signal map first; if it yields
    nothing and the register is CAN, the variable-name override; else None.
    Curated and conservative — an unrecognized signal stays unresolved rather
    than guessing (see `_CLASSIFICATION_SIGNALS`).
    """
    by_signal = _resolve_by_signal(external_classification)
    if by_signal is not None:
        return by_signal
    if abbrev == "can":
        return _CAN_VARNAME_CLASSIFICATION.get(var_name)
    return None


# Slug stem for the synthesized variant of variant-less registers (LSS/BU/SOL).
_DEFAULT_VARIANT = "_default"

_FILENAME_ABBREV_RE = re.compile(r"\(([A-Za-zÅÄÖåäö0-9_]{2,10})\)")


def _sos_abbrev(reg: SosRegister) -> str:
    """The stable short token for `mint("sos", <abbrev>, …)`.

    Sourced from the parenthesized code in the workbook FILENAME stem
    ("...(LMED)_webb.xlsx" -> "lmed"); R4-verified unique across all 13
    workbooks. Lowercased so the mint token and the allow-list keys share one
    casing convention. Falls back to a normalized `dataset_name` only if the
    filename carries no parenthesized code (none of the 13 deliveries hit the
    fallback) — keeps the build from minting a register id off an empty token.
    """
    m = _FILENAME_ABBREV_RE.search(reg.source_file.stem)
    if m:
        return m.group(1).lower()
    base = (reg.dataset_name or reg.source_file.stem).strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", base).strip("_") or "sos_register"


def _norm_data_type(dt: str | None) -> str | None:
    """Lowercase/trim the SOS `Datatyp` to the canonical comparison form.

    Used both as the split-seam discriminator and to mirror the nullable
    `variable_state.data_type` column. SOS labels are Swedish free text
    ("Datum", "Heltal", "Sträng (text)"); identity is by normalized string.
    """
    if dt is None:
        return None
    s = dt.strip().lower()
    return s or None


def _shape_signature(group: list[SosVariable], codes: tuple[str, ...]) -> str:
    """Stable shape-derived discriminator for a split sibling (NOT a counter).

    Combines the normalized data_type with a hash of the sorted code-set so
    rebuilds mint byte-identical sibling ids. The known SOS splits are
    data_type-only (empty code-set), so data_type alone discriminates them; the
    code-set component future-proofs a disjoint-codelist split.
    """
    dts = sorted({_norm_data_type(v.data_type) or "" for v in group})
    h = _value_set_hash([(c, "") for c in codes]) if codes else b""
    return "|".join(dts) + ":" + h.hex()[:16]


def _parse_tidsperiod(tp: str | None) -> tuple[str | None, str | None]:
    """Parse a SOS `Tidsperiod` token into ISO 8601 [from, to] bounds.

    Forms (consumed verbatim from the parser's forward-filled value):
      "YYYY-YYYY" -> [YYYY-01-01, YYYY-12-31]
      "YYYY-"     -> [YYYY-01-01, None]   (open-ended)
      "YYYY"      -> [YYYY-01-01, YYYY-12-31]
      None / unparseable -> (None, None)  (unbounded; caller falls back to
                                            deldatamängd / variable era)
    """
    if tp is None:
        return None, None
    s = tp.strip()
    m = re.fullmatch(r"(\d{4})\s*-\s*(\d{4})", s)
    if m:
        return f"{m.group(1)}-01-01", f"{m.group(2)}-12-31"
    m = re.fullmatch(r"(\d{4})\s*-\s*", s)
    if m:
        return f"{m.group(1)}-01-01", None
    m = re.fullmatch(r"(\d{4})", s)
    if m:
        return f"{m.group(1)}-01-01", f"{m.group(1)}-12-31"
    return None, None


# Code charset for a `Värdemängd` enumeration: digits + Latin/Swedish letters
# plus `.`/`_`/`-` (no whitespace, comma, or colon). A bare `-` separating two
# digits (`0-744`) is a NUMERIC RANGE descriptor, not an enumeration, so it is
# rejected even though `-` is otherwise an allowed code char.
_VALUE_CODE_CHARSET = re.compile(r"[0-9A-Za-zÅÄÖåäö._-]+")


def _clean_value_code(c: str) -> str | None:
    """Validate one `Värdemängd` code token; return it stripped, or ``None`` if
    it isn't a clean enumeration code.

    Rejects: empty; embedded whitespace/comma/colon (label/prose leakage);
    a numeric range (`0-744`); anything outside `_VALUE_CODE_CHARSET`. Only the
    CODE is constrained this tightly — labels are free-form (see the classifier).
    """
    c = c.strip()
    if not c:
        return None
    if any(ch in c for ch in (" ", "\t", ",", ":")):
        return None
    if re.search(r"\d-\d", c):  # numeric range like 0-744, not a code
        return None
    if not _VALUE_CODE_CHARSET.fullmatch(c):
        return None
    return c


def _classify_value_set_text(text: str | None) -> list[tuple[str, str | None]] | None:
    """Classify a raw SOS `Värdemängd` cell into (code, label) pairs, or reject.

    This is the #401 fallback that promotes a variable's INLINE enumerated code
    list to a value set when the variable has no `Kodlista_*` sheet (the #373
    deferral — styrtabell decode tables were excluded from minting, but their
    `Värdemängd` enumeration was never bound). It is deliberately CONSERVATIVE:
    rejecting (returning ``None``) leaves the variable exactly as today (no value
    set), so a wrong reject is a no-op while a wrong ACCEPT would mint garbage.

    Two accepted forms (real-corpus-verified):
      - `kod=klartext` pairs — every segment carries `=`: code with inline label
        (`1=ja; 0=nej`, newline-delimited lists). Label is the right of the first
        `=` (labels may contain spaces/commas/colons — only the code is checked).
      - bare codes — no segment carries `=`: code-only (`1;2;3;4;5;9`, `LEG;SPEC`).

    Rejected (free-text trap): single segment (a descriptor like `Fritext`);
    MIXED `=`/no-`=` (catches trailing-prose cells like `0=…; …; strängen är
    tom`); any invalid code (range, comma, colon, whitespace); duplicate codes.
    """
    if not text or not text.strip():
        return None
    # Split on `;` AND newline simultaneously — both are clean SOS separators.
    segments = [s.strip() for s in re.split(r"[;\n]+", text) if s.strip()]
    # A single segment is a free-text descriptor, not an enumeration.
    if len(segments) < 2:
        return None

    with_eq = sum(1 for s in segments if "=" in s)
    if with_eq == len(segments):
        # kod=klartext: partition each on the FIRST `=`.
        pairs: list[tuple[str, str | None]] = []
        for s in segments:
            raw_code, _, raw_label = s.partition("=")
            code = _clean_value_code(raw_code)
            label = raw_label.strip()
            if code is None or not label:
                return None
            pairs.append((code, label))
    elif with_eq == 0:
        # bare codes: each segment IS the code, no label.
        pairs = []
        for s in segments:
            code = _clean_value_code(s)
            if code is None:
                return None
            pairs.append((code, None))
    else:
        # MIXED `=`/no-`=` -> trailing-prose / malformed cell. Reject.
        return None

    if len({code for code, _ in pairs}) != len(pairs):  # duplicate codes -> reject
        return None
    return pairs


def _iso_bound(value: int | None, *, end: bool) -> str | None:
    """Normalize a SOS `data_från`/`data_till` int to a full ISO date.

    SOS deldatamängd/variable date bounds arrive in three shapes (R4-verified):
    `YYYY` (4 digits), `YYYYMM` (6), `YYYYMMDD` (8). Expand the coarse forms to a
    full-date range bound (the `variable_state` CHECK requires length-10 ISO):
      4 digits: from -> YYYY-01-01, to -> YYYY-12-31
      6 digits: from -> YYYY-MM-01, to -> YYYY-MM-<last-day>
      8 digits: exact YYYY-MM-DD (both ends)
    Returns ``None`` for ``None`` or an unrecognized shape (treated as unbounded).
    """
    if value is None:
        return None
    s = str(value)
    if len(s) == 4 and s.isdigit():
        return f"{s}-12-31" if end else f"{s}-01-01"
    if len(s) == 6 and s.isdigit():
        year, month = int(s[:4]), int(s[4:6])
        if not (1 <= month <= 12):
            return None
        if end:
            last = calendar.monthrange(year, month)[1]
            return f"{year:04d}-{month:02d}-{last:02d}"
        return f"{year:04d}-{month:02d}-01"
    if len(s) == 8 and s.isdigit():
        year, month, day = int(s[:4]), int(s[4:6]), int(s[6:8])
        if not (1 <= month <= 12 and 1 <= day <= 31):
            return None
        return f"{year:04d}-{month:02d}-{day:02d}"
    return None


def _intersect_window(
    bounds: list[tuple[str | None, str | None]],
) -> tuple[str | None, str | None] | None:
    """3-way validity intersection: [max(froms) .. min(tos)] over the given
    (from, to) bounds. ``None`` froms are -inf, ``None`` tos are +inf. Returns
    the intersected (from, to), or ``None`` if the window is empty (from > to).
    The full-date contract makes ISO strings lexically == chronologically
    comparable.
    """
    lo: str | None = None
    for f, _t in bounds:
        if f is not None and (lo is None or f > lo):
            lo = f
    hi: str | None = None
    for _f, t in bounds:
        if t is not None and (hi is None or t < hi):
            hi = t
    if lo is not None and hi is not None and lo > hi:
        return None
    return lo, hi


def _intersect_advisory_deldat(
    authoritative: list[tuple[str | None, str | None]],
    deldat: tuple[str | None, str | None],
) -> tuple[tuple[str | None, str | None] | None, bool]:
    """P2#2: intersect ``authoritative`` (the variable window + any code window)
    with the ADVISORY deldatamängd coarse window ``deldat``.

    The deldatamängd data_from/to is an advisory coarse bound; the VARIABLE (and
    code, for code-bearing states) window is authoritative. If including the
    deldat bound makes the window empty BUT the authoritative bounds alone are
    non-empty, DROP the deldat bound and keep the authoritative window. The code
    tidsperiod (folded into ``authoritative`` by the caller) stays a real bound —
    only the deldat bound is the one that yields.

    Returns ``(window, deldat_dropped)``: ``window`` is ``None`` only when the
    authoritative bounds alone are empty (a genuine empty state, dropped as
    before); ``deldat_dropped`` is ``True`` when the deldat bound was contradicted
    and discarded (the caller emits an IRWarning).
    """
    full = _intersect_window([*authoritative, deldat])
    if full is not None:
        return full, False
    # deldat made it empty — fall back to the authoritative window alone.
    auth = _intersect_window(authoritative)
    if auth is None:
        return None, False  # genuinely empty even without deldat
    return auth, True


def _widest_valid_to(a: str | None, b: str | None) -> str | None:
    """The wider of two ``valid_to`` bounds: ``None`` is open-ended (the widest
    possible), else the later ISO date (full-date strings sort chronologically).
    Reconciles merged-member states that share a (variant, valid_from) state_id
    but carry different end bounds — see ``_emit_states``."""
    if a is None or b is None:
        return None
    return max(a, b)


def _windows_overlap(
    a_from: str | None, a_to: str | None, b_from: str | None, b_to: str | None
) -> bool:
    """Closed-interval overlap test, mirroring the build invariant's SQL
    (``validate.py`` → "one value_set per (variable, variant, period, column)")
    and ``catalog._states_in_bounds``: ``a_from <= b_to AND b_from <= a_to``.
    A ``None`` ``valid_from`` is open-start (−∞); a ``None`` ``valid_to`` is
    open-end (+∞). Full-date ISO strings sort chronologically, so once the open
    bounds are substituted the comparison is a plain string compare."""
    lo = "0001-01-01"  # −∞ (open-start)
    hi = "9999-12-31"  # +∞ (open-end); the sentinel the validator compares to
    af, at_ = a_from or lo, a_to or hi
    bf, bt = b_from or lo, b_to or hi
    return af <= bt and bf <= at_


# Deferred-write identity for a #401 Värdemängd value set (see `_emit_states`):
# (member_hash, codes). `member_hash` is the SAME content-addressed digest
# `_ensure_value_set` computes — over `[(c.code, c.label) for c in codes]` — so a
# deferred write content-SHARES identically; `codes` is the IRValueCode list the
# survivor write replays into `_ensure_value_set`. `None` marks a code-less state
# (free-text Värdemängd, or a binding suppressed by reconciliation). Only the
# Värdemängd path defers; kodlista + entity-registry paths still write eagerly
# (they segment/collapse so they can never orphan).
PendingValueSet = tuple[bytes, list[IRValueCode]]


# Sentinel for `collect`'s pending-identity arg meaning "eager path, no #464
# deferral": the kodlista + entity-registry callers pass no identity, distinct
# from the Värdemängd path passing `None` (an explicit code-less deferral). The
# sentinel keeps the Värdemängd `None` from being mistaken for "untracked".
class _NoPending:
    __slots__ = ()


_NO_PENDING = _NoPending()

# `collect`'s signature: (state, pending) where `pending` defaults to the eager
# sentinel so kodlista-path callers keep calling `collect(obj)` with one arg. A
# Protocol (not a `Callable[...]` alias) is required to express the optional
# second parameter — `Callable` can't carry a default.
if TYPE_CHECKING:

    class _CollectFn(Protocol):
        def __call__(
            self,
            obj: IRObject,
            pending: PendingValueSet | None | _NoPending = ...,
        ) -> Iterator[IRObject]: ...


def _value_code(
    code: str, label: str | None, window: tuple[str | None, str | None]
) -> IRValueCode:
    """Build a value-set IRValueCode with placeholder ids (code_id/value_set_id
    are assigned at write-back by `_ensure_value_set`). `window` =
    (valid_from, valid_to)."""
    return IRValueCode(
        code_id=0,
        value_set_id=0,
        code=code,
        label=label or "",
        valid_from=window[0],
        valid_to=window[1],
    )


_SEG_LO = date(1, 1, 1)  # open-start sentinel (−∞)
# Open-end sentinel (+∞). A real closed window literally ending 9999-12-31 would
# be conflated with open-ended and emitted as `None`, but `_parse_tidsperiod`
# never produces a literal 9999 bound (open windows yield `None`), so this only
# bites a pathological year-9999 tidsperiod — accepted.
_SEG_HI = date(9999, 12, 31)


def _segment_windowed_codes(
    windowed: list[tuple[str | None, str | None, IRValueCode]],
) -> list[tuple[str | None, str | None, list[IRValueCode]]]:
    """Partition heterogeneous per-code validity windows into NON-OVERLAPPING
    period segments (sweep-line), each carrying the UNION of codes live across its
    full extent — so a period resolves to exactly ONE value set on the column
    (see DESIGN.md → Build-time triage (SCB)). This is the deferred "Path B" per-period refinement: it
    replaces bucketing codes by their EXACT window (which produced overlapping
    value sets when a wide/open code coexisted with narrower sub-windows, e.g.
    SOS ALKOHOL `'0'`[1987–] over `'1'`[1987–96]/[1997–]).

    Windows are closed ISO-date intervals; `None` bounds are open (−∞ / +∞). Cuts
    fall at every window start and the day AFTER every window end. Adjacent
    segments with an identical `(code, label)` union are merged (RLE) so an
    unchanged codelist doesn't fragment. Returns `(valid_from, valid_to, codes)`
    in chronological order; `None` bounds stay open.
    """
    if not windowed:
        return []

    def _lo(s: str | None) -> date:
        return date.fromisoformat(s) if s else _SEG_LO

    def _hi(s: str | None) -> date:
        return date.fromisoformat(s) if s else _SEG_HI

    intervals = [(_lo(vf), _hi(vt), code) for vf, vt, code in windowed]
    # Segment starts: every window start + the day after every (closed) window end.
    cuts: set[date] = set()
    for lo, hi, _ in intervals:
        cuts.add(lo)
        if hi < _SEG_HI:
            cuts.add(hi + timedelta(days=1))
    starts = sorted(cuts)

    # raw: list of (lo, hi, codes) — non-overlapping; skip stretches no code
    # covers. The per-segment rescan of `intervals` is O(segments × intervals),
    # which is negligible at SOS kodlista sizes. Each segment's live codes are
    # DEDUPED by (code, label): two identical rows with overlapping tidsperiod
    # windows would otherwise both land here, and `_ensure_value_set` hashes the
    # pair list while `value_set_member`'s PK collapses the duplicate — desyncing
    # `member_hash` from the stored member set and breaking content-share.
    raw: list[tuple[date, date, list[IRValueCode]]] = []
    for i, seg_lo in enumerate(starts):
        seg_hi = starts[i + 1] - timedelta(days=1) if i + 1 < len(starts) else _SEG_HI
        seen: set[tuple[str, str]] = set()
        live: list[IRValueCode] = []
        for lo, hi, c in intervals:
            if lo <= seg_lo and seg_hi <= hi and (c.code, c.label) not in seen:
                seen.add((c.code, c.label))
                live.append(c)
        if live:
            raw.append((seg_lo, seg_hi, live))

    # RLE-merge CONTIGUOUS segments whose (code, label) union is identical.
    merged: list[tuple[date, date, list[IRValueCode], frozenset[tuple[str, str]]]] = []
    for seg_lo, seg_hi, live in raw:
        sig = frozenset((c.code, c.label) for c in live)
        if merged:
            plo, phi, plive, psig = merged[-1]
            if psig == sig and phi < _SEG_HI and seg_lo == phi + timedelta(days=1):
                merged[-1] = (plo, seg_hi, plive, sig)
                continue
        merged.append((seg_lo, seg_hi, live, sig))

    return [
        (
            None if lo == _SEG_LO else lo.isoformat(),
            None if hi == _SEG_HI else hi.isoformat(),
            live,
        )
        for lo, hi, live, _sig in merged
    ]


class SOSAdapter:
    """Socialstyrelsen source adapter (IRAdapter).

    `emit(source_dir)` parses every `.xlsx` under ``source_dir`` (the
    `Socialstyrelsen/` directory) into `SosRegister` trees and yields the
    provider-neutral IR stream the materializer consumes. PURELY ADDITIVE:
    `mint()`ed ids (band `[2^62, 2^63)`), content-shared value_sets.

    Mirrors the `SCBAdapter` attribute surface `materialize()` reads:
      - `source_checksums`, `row_counts` — manifest inputs.
      - `coalesce_stats` — minimal SOS analog (`{}` shape ok).
      - `fold_slug_hints` — SOS folds are rare; left empty.
      - `sibling_edges` — split-sibling (variable_id, variable_id) pairs the
        concept-group edge fold reads directly (`edge_siblings`); build-only
        routing, not persisted to any shipped table.
    """

    provider = "sos"

    def __init__(self, conn: Any) -> None:
        # The connection is bound for API parity with SCBAdapter and so SOS can
        # consult/dedup value tables (INSERT OR IGNORE + read-back) against the
        # rows SCB already wrote. SOS writes NO build-scratch.
        self.conn = conn
        self.source_checksums: dict[str, str] = {}
        self.row_counts: dict[str, int] = {}
        self.coalesce_stats: dict[str, Any] = {}
        self.fold_slug_hints: dict[int, str] = {}
        self.sibling_edges: list[tuple[int, int]] = []
        # Provider-blind classification linkage: (variable_id, value_set_id,
        # short_name) per variable_state of a RESOLVED variable. `db.py` resolves
        # short_name → classification_id and feeds `classification_candidate`,
        # which `_backfill_state_classifications` reads — same path as SCB.
        # value_set_id is often None (code-less ICD/ATC states); the backfill
        # keys on (variable_id, value_set_id) with `IS`, so NULL is fine.
        self.classification_candidates: list[tuple[int, int | None, str]] = []
        # Maintainer visibility (the issue wants this): distinct
        # external_classification values resolved vs unresolved, + tagged counts.
        self._resolved_ext: set[str] = set()
        self._unresolved_ext: set[str] = set()
        self._resolved_variables = 0
        self._tagged_states = 0
        # Content-share cache: member_hash -> value_set_id (read back from the DB
        # so an identical SOS code list collapses onto SCB's existing row).
        self._set_id_by_hash: dict[bytes, int] = {}

    # -- value-table hybrid (R2) -------------------------------------------

    def _ensure_value_set(self, codes: list[IRValueCode]) -> int | None:
        """Write a SOS value_set (+codes+members) via INSERT OR IGNORE +
        read-back, so it content-SHARES any identical SCB/SOS row.

        `value_code` dedups on UNIQUE(code, label); `value_set` dedups on
        UNIQUE(member_hash). Both ids stay AUTOINCREMENT, content-addressed,
        provider-shared (unbanded — excluded from the band assertion).
        Returns the shared value_set_id, or ``None`` for an empty code list.

        Ordering is load-bearing: the materializer runs SCB before SOS, so by
        the time this writes, SCB's explicit-counter `code_id`s have set the
        `value_code` rowid high-water mark and these INSERT-OR-IGNOREs pick up
        after them without collision. Only `code`/`label` are read here; each
        `IRValueCode.valid_from`/`valid_to` is used upstream to decide code
        survival but not persisted — `value_code` has no validity columns
        (per-code validity awaits the materializer-owned value tables, Path B).
        """
        if not codes:
            return None
        conn = self.conn
        pairs = [(c.code, c.label) for c in codes]
        member_hash = _value_set_hash(pairs)
        cached = self._set_id_by_hash.get(member_hash)
        if cached is not None:
            # Same content hash already written — codes + members exist.
            return cached
        # `value_code`: one row per distinct (code, label); read back its id.
        code_id_of: dict[tuple[str, str], int] = {}
        for code, label in pairs:
            conn.execute(
                "INSERT OR IGNORE INTO value_code (code, label) VALUES (?, ?)",
                (code, label),
            )
            row = conn.execute(
                "SELECT code_id FROM value_code WHERE code = ? AND label = ?",
                (code, label),
            ).fetchone()
            code_id_of[(code, label)] = row[0]
        conn.execute(
            "INSERT OR IGNORE INTO value_set (member_hash) VALUES (?)", (member_hash,)
        )
        set_id = conn.execute(
            "SELECT value_set_id FROM value_set WHERE member_hash = ?", (member_hash,)
        ).fetchone()[0]
        for code, label in pairs:
            conn.execute(
                "INSERT OR IGNORE INTO value_set_member (value_set_id, code_id) "
                "VALUES (?, ?)",
                (set_id, code_id_of[(code, label)]),
            )
        self._set_id_by_hash[member_hash] = set_id
        return set_id

    # -- emit ---------------------------------------------------------------

    def emit(self, source_dir: Path) -> Iterator[IRObject]:
        """Parse the `Socialstyrelsen/` workbooks under ``source_dir`` and emit
        IR objects in FK-topological order (register -> variant -> value_set ->
        variable -> state/alias -> related-to -> provenance/warnings).

        Unreadable workbooks become an `IRWarning` (not a build abort): each
        `.xlsx` parses in its own try/except so one corrupt delivery can't sink
        the others.
        """
        for path in sorted(source_dir.iterdir()):
            if not path.is_file():
                continue
            if path.suffix.lower() != ".xlsx" or path.name.startswith("~$"):
                continue
            self.source_checksums[path.name] = _file_sha256(path)
            try:
                reg = parse_register_file(path)
            except SosParseError as exc:
                # Unreadable workbook: surface, skip. entity_id 0 (no register
                # minted yet). The provenance DB records it; the build proceeds.
                yield IRWarning(
                    entity_kind="register",
                    entity_id=0,
                    code="sos_workbook_unreadable",
                    detail=f"{path.name}: {exc}",
                )
                continue
            yield from self._emit_register(reg)

    def _emit_register(self, reg: SosRegister) -> Iterator[IRObject]:
        abbrev = _sos_abbrev(reg)
        register_id = mint("sos", abbrev)
        self.row_counts[f"sos:{abbrev}"] = len(reg.variables)

        name = reg.dcat_ap.title_sv or reg.dataset_name or abbrev.upper()
        yield IRRegister(
            register_id=register_id,
            provider="sos",
            slug="",  # ignored by the reinsert; populate_slugs fills it
            name=name,
            description=reg.dcat_ap.description_sv,
            purpose=reg.dcat_ap.description_sv,
        )

        # Surface every parser warning (missing sheets, unparseable kodlistor)
        # as an IRWarning keyed on the register.
        for w in reg.warnings:
            yield IRWarning(
                entity_kind="register",
                entity_id=register_id,
                code="sos_parser_warning",
                detail=w,
            )

        # -- styrtabell detection (#373): decode tables (kod->klartext) are
        # EXCLUDED from variant + variable minting — their rows bind to the
        # coded variable's value set, not to a research variable. A styrtabell
        # is identified by BOTH structural signals agreeing (`_is_styrtabell`).
        # When exactly one signal is present (XOR) the shape has drifted — warn
        # so the half-flagged table surfaces instead of silently minting; such a
        # deldatamängd is NOT excluded (it stays in variant/variable minting).
        styr_names = {d.name for d in reg.deldatamangder if _is_styrtabell(d)}
        for d in reg.deldatamangder:
            agg = (d.aggregation_level or "").strip().casefold() == "ej relevant"
            etikett = (d.label or "").strip().casefold().startswith("styrtabell")
            if agg != etikett:
                present, missing = (
                    ("Aggregeringsnivå='Ej relevant'", "Deldatamängdsetikett")
                    if agg
                    else ("Deldatamängdsetikett='Styrtabell …'", "Aggregeringsnivå")
                )
                yield IRWarning(
                    entity_kind="register",
                    entity_id=register_id,
                    code="sos_styrtabell_signal_mismatch",
                    detail=(
                        f"{abbrev}/{d.name}: styrtabell signal {present} present "
                        f"but {missing} signal missing; not excluded — check "
                        "whether this is a decode table (#373)"
                    ),
                )

        # -- variant synthesis: detect via DELDATAMÄNGDER-SHEET ABSENCE
        # (r.deldatamangder == ()), NOT var.deldatamangd (BU populates that from
        # a Datavynamn column yet has no Deldatamängder sheet -> variant-less).
        variant_less = reg.deldatamangder == ()
        variant_id_of: dict[str | None, int] = {}
        if variant_less:
            default_variant_id = mint("sos", abbrev, _DEFAULT_VARIANT)
            yield IRVariant(
                register_variant_id=default_variant_id,
                register_id=register_id,
                slug="",
                name=_DEFAULT_VARIANT,
                description=None,
                synthesized=True,
            )
            variant_id_of[None] = default_variant_id
        else:
            # Dedup deldatamängder by name: a workbook can carry the same name
            # twice (e.g. LOVA ships two 'LOVA' rows) — they mint the same
            # variant id, so emit ONE IRVariant (first occurrence wins) or the
            # reinsert hits a register_variant PK collision.
            for d in reg.deldatamangder:
                # styrtabell decode tables (#373): never minted as a variant —
                # skip before variant_id_of/provenance/deldat_by_name see them.
                if d.name in styr_names:
                    continue
                if d.name in variant_id_of:
                    continue
                vid = mint("sos", abbrev, d.name)
                variant_id_of[d.name] = vid
                yield IRVariant(
                    register_variant_id=vid,
                    register_id=register_id,
                    slug="",
                    name=d.name,
                    description=d.description or d.label,
                )

        # Advisory-window source per deldatamängd name. A name carried by >1
        # sheet row (LOVA ships two 'LOVA' rows) is AMBIGUOUS: the rows mint ONE
        # variant but may carry different data_från/till windows, and there is
        # no curated token<->row pairing — so an ambiguous name contributes NO
        # deldat window (None; the authoritative variable window still applies).
        name_counts = Counter(d.name for d in reg.deldatamangder)
        deldat_by_name = {
            d.name: d for d in reg.deldatamangder if name_counts[d.name] == 1
        }

        # -- kodlista resolution: variable_hint -> kodlista (case-insensitive).
        # Skip unparseable kodlistor (raw_rows non-empty) for value-set
        # construction; warn instead of fabricating.
        # `raw_kodlista_hints` records hints that HAD a kodlista sheet but were
        # skipped as raw: kodlista-wins means such a variable must stay code-less
        # (`kodlista is None` below) rather than fall back to fabricating inline
        # codes from its Värdemängd (#401 Fix B; defensive — 0 corpus occurrences).
        kodlista_by_var: dict[str, SosKodlista] = {}
        raw_kodlista_hints: set[str] = set()
        for k in reg.kodlistor:
            if k.raw_rows:
                raw_kodlista_hints.add(k.variable_hint.lower())
                yield IRWarning(
                    entity_kind="register",
                    entity_id=register_id,
                    code="sos_kodlista_unparseable",
                    detail=f"{k.sheet_name}: no Tidsperiod/Kod header; skipped",
                )
                continue
            kodlista_by_var[k.variable_hint.lower()] = k

        # -- group variables by name (the merge key + provider_key). Drop
        # styrtabell decode-column rows first (#373): a row whose deldatamängd
        # token resolves ENTIRELY to styrtabell names (same resolution the
        # member loop uses) carries decode-only columns (KLARTEXT/BESKRIVNING …)
        # bound to a value set, not a research variable. A coded variable that
        # ALSO appears under a real deldatamängd survives via its other rows.
        def _is_styrtabell_var(v: SosVariable) -> bool:
            targets = DELDATAMANGD_TOKEN_MAP.get(
                (abbrev, v.deldatamangd or ""), (v.deldatamangd,)
            )
            return bool(targets) and all(t in styr_names for t in targets)

        groups: dict[str, list[SosVariable]] = {}
        for v in reg.variables:
            if _is_styrtabell_var(v):
                continue
            # Apply curated single-row name corrections (#362) before grouping,
            # so a mistyped duplicate-name row keys into its OWN group instead of
            # silently merging. Rewrite the MEMBER's name (not just the group key):
            # for SOS the delivery column == the variable name, so the corrected
            # name must flow through to col/delivery_column_name in
            # _emit_member_states — otherwise the de-merged variable would emit its
            # alias/state with the WRONG (mistyped) column, colliding with the real
            # variable that legitimately owns it. The corrected name then becomes
            # the group key -> provider_key/mint/slug in _emit_merged.
            corrected = VARIABLE_NAME_CORRECTIONS.get((abbrev, v.name, v.label or ""))
            if corrected is not None:
                v = replace(v, name=corrected)
            groups.setdefault(v.name, []).append(v)

        for name_key in sorted(groups):
            group = groups[name_key]
            yield from self._emit_variable_group(
                abbrev=abbrev,
                register_id=register_id,
                name=name_key,
                group=group,
                variant_id_of=variant_id_of,
                variant_less=variant_less,
                deldat_by_name=deldat_by_name,
                kodlista=kodlista_by_var.get(name_key.lower()),
                has_kodlista_sheet=name_key.lower() in raw_kodlista_hints,
            )

        # -- delivery provenance, one per variant. SOS carries no approval
        # tokens, so emit_when_no_tokens records one bare-token row per variant.
        for vid in variant_id_of.values():
            yield IRDeliveryProvenance(
                register_id=register_id,
                register_variant_id=vid,
                source_file=reg.source_file.name,
                delivery_version=reg.dataset_version,
                delivery_date=reg.dataset_date,
                template_version=reg.template_version,
                emit_when_no_tokens=True,
            )

    def _emit_variable_group(
        self,
        *,
        abbrev: str,
        register_id: int,
        name: str,
        group: list[SosVariable],
        variant_id_of: dict[str | None, int],
        variant_less: bool,
        deldat_by_name: dict[str, SosDeldatamangd],
        kodlista: SosKodlista | None,
        has_kodlista_sheet: bool,
    ) -> Iterator[IRObject]:
        """Emit one variable group: MERGE (default), SPLIT (allow-list), or
        WARN-MERGE (unanticipated conflict). Then emit its era-windowed states.

        `has_kodlista_sheet`: this variable HAD a kodlista sheet that was skipped
        as unparseable (`raw_rows`) — kodlista-wins, so its Värdemängd fallback is
        suppressed downstream (stays code-less) instead of fabricating codes.
        """
        # -- conflict detection: incompatible normalized data_type OR disjoint
        # structured code-list shape (NEVER value_set_text — the free-text trap).
        data_types = {_norm_data_type(v.data_type) for v in group}
        conflict = len(group) > 1 and len(data_types) > 1
        key = (abbrev, name)

        if conflict and key in KNOWN_SPLIT_ALLOWLIST:
            yield from self._emit_split(
                abbrev=abbrev,
                register_id=register_id,
                name=name,
                group=group,
                variant_id_of=variant_id_of,
                variant_less=variant_less,
                deldat_by_name=deldat_by_name,
                kodlista=kodlista,
                has_kodlista_sheet=has_kodlista_sheet,
            )
            return

        if conflict and key not in KNOWN_MERGE_ALLOWLIST:
            # Fail-soft: unanticipated same-name conflict -> WARN + MERGE.
            # Allow-listed keys (#362) still merge below; only the warn is
            # silenced — the merge is intentional and type-lossless.
            yield IRWarning(
                entity_kind="variable",
                entity_id=mint("sos", abbrev, name),
                code="sos_unanticipated_same_name_conflict",
                detail=(
                    f"{abbrev}/{name}: data_type diverges "
                    f"{sorted(dt or '' for dt in data_types)}; merged"
                ),
            )

        yield from self._emit_merged(
            abbrev=abbrev,
            register_id=register_id,
            name=name,
            group=group,
            variant_id_of=variant_id_of,
            variant_less=variant_less,
            deldat_by_name=deldat_by_name,
            kodlista=kodlista,
            has_kodlista_sheet=has_kodlista_sheet,
        )

    def _emit_merged(
        self,
        *,
        abbrev: str,
        register_id: int,
        name: str,
        group: list[SosVariable],
        variant_id_of: dict[str | None, int],
        variant_less: bool,
        deldat_by_name: dict[str, SosDeldatamangd],
        kodlista: SosKodlista | None,
        has_kodlista_sheet: bool,
    ) -> Iterator[IRObject]:
        variable_id = mint("sos", abbrev, name)
        # Deterministic winner for scalar fields: first non-null across the
        # group (group is delivery-order stable per the parser).
        label = _first(group, "label")
        description = _first(group, "description")
        yield IRVariable(
            variable_id=variable_id,
            register_id=register_id,
            provider_key=name,
            slug="",
            name=label or name,
            definition=None,
            description=description,
            measurement_unit=None,
            is_sensitive=False,
            is_identifier=False,
            source_register_id=None,
            source_register_text=None,
            source_label=None,
        )
        classification = self._resolve_group_classification(abbrev, name, group)
        yield from self._emit_states(
            abbrev=abbrev,
            variable_id=variable_id,
            members=group,
            variant_id_of=variant_id_of,
            variant_less=variant_less,
            deldat_by_name=deldat_by_name,
            kodlista=kodlista,
            entity_registry=(abbrev, kodlista.variable_hint)
            in ENTITY_REGISTRY_KODLISTOR
            if kodlista
            else False,
            classification=classification,
            has_kodlista_sheet=has_kodlista_sheet,
        )

    def _emit_split(
        self,
        *,
        abbrev: str,
        register_id: int,
        name: str,
        group: list[SosVariable],
        variant_id_of: dict[str | None, int],
        variant_less: bool,
        deldat_by_name: dict[str, SosDeldatamangd],
        kodlista: SosKodlista | None,
        has_kodlista_sheet: bool,
    ) -> Iterator[IRObject]:
        # Partition the group by normalized data_type seam (the known SOS splits
        # are data_type-only, no codelist). One sibling per distinct shape.
        by_shape: dict[str | None, list[SosVariable]] = {}
        for v in group:
            by_shape.setdefault(_norm_data_type(v.data_type), []).append(v)
        sibling_ids: list[int] = []
        for shape in sorted(by_shape, key=lambda s: s or ""):
            members = by_shape[shape]
            disc = _shape_signature(members, ())
            variable_id = mint("sos", abbrev, name, disc)
            sibling_ids.append(variable_id)
            label = _first(members, "label")
            description = _first(members, "description")
            yield IRVariable(
                variable_id=variable_id,
                register_id=register_id,
                provider_key=name,  # split siblings SHARE the source name
                slug="",
                name=label or name,
                definition=None,
                description=description,
                measurement_unit=None,
                is_sensitive=False,
                is_identifier=False,
                source_register_id=None,
                source_register_text=None,
                source_label=None,
            )
            classification = self._resolve_group_classification(abbrev, name, members)
            yield from self._emit_states(
                abbrev=abbrev,
                variable_id=variable_id,
                members=members,
                variant_id_of=variant_id_of,
                variant_less=variant_less,
                deldat_by_name=deldat_by_name,
                kodlista=kodlista,
                entity_registry=False,
                classification=classification,
                has_kodlista_sheet=has_kodlista_sheet,
            )
        # (N choose 2) sibling pairs -> self.sibling_edges, feeding the
        # concept-group edge fold (`edge_siblings`). Emit each unordered pair once.
        for i in range(len(sibling_ids)):
            for j in range(i + 1, len(sibling_ids)):
                self.sibling_edges.append((sibling_ids[i], sibling_ids[j]))

    def _resolve_group_classification(
        self, abbrev: str, name: str, group: list[SosVariable]
    ) -> str | None:
        """Resolve one variable group → classification short_name (or None) and
        record the distinct external_classification value as resolved/unresolved
        for the maintainer summary. The signal is the first non-null
        `external_classification` across the group (delivery-order stable).

        The resolved/unresolved split tracks the SIGNAL-MAP outcome only: a CAN
        variable-name override resolves the variable but its ext value names no
        seeded system, so it stays in `_unresolved_ext` (the same ext value can
        also back a sibling that the override leaves None — counting it resolved
        would misreport the resolver's URL-signal reach)."""
        ext = _first(group, "external_classification")
        short_name = _resolve_classification(ext, abbrev, name)
        if ext:
            by_signal = _resolve_by_signal(ext)
            (self._resolved_ext if by_signal else self._unresolved_ext).add(ext)
        return short_name

    def classification_summary(self) -> str:
        """One-line maintainer summary of the classification resolver's reach
        (distinct external_classification values resolved vs unresolved, plus
        variables/states tagged). Logged by `db.py` after the SOS feed runs."""
        return (
            f"SOS classification resolver: "
            f"{len(self._resolved_ext)} distinct external_classification values "
            f"resolved, {len(self._unresolved_ext)} unresolved; "
            f"{self._resolved_variables} variables / {self._tagged_states} states "
            f"tagged ({len(self.classification_candidates)} candidates)"
        )

    def _emit_states(
        self,
        *,
        abbrev: str,
        variable_id: int,
        members: list[SosVariable],
        variant_id_of: dict[str | None, int],
        variant_less: bool,
        deldat_by_name: dict[str, SosDeldatamangd],
        kodlista: SosKodlista | None,
        entity_registry: bool,
        classification: str | None,
        has_kodlista_sheet: bool = False,
    ) -> Iterator[IRObject]:
        """Emit IRVariableState(+IRValueSet/IRValueCode) + IRVariableAlias for
        each (member, resolved-era windowed value-set) of a variable.

        Two members of a MERGED variable can resolve to the same variant + same
        `valid_from` (e.g. a variable name appearing under a duplicate
        deldatamängd, or members whose windows share a start) -> the same minted
        state_id, because the `idx_variable_state_unique` basis EXCLUDES
        `valid_to`. They are NOT duplicates: each carries its member's own
        `valid_to`.

        `valid_to` + `value_set_id` reconciliation. The divergent-window /
        prefer-coded reconciliation below is VÄRDEMÄNGD-ONLY (`kodlista is None`):
        a kodlista-derived collision (entity-registry or windowed) keeps the
        ORIGINAL pre-#401 behavior — always widen `valid_to`, keep `prior` (incl.
        its value_set_id) — because a distinguishing kodlista code buckets into a
        non-None `value_set_version_label` that changes the state_id, so a
        same-state_id kodlista collision provably carries the SAME value set.

        For the Värdemängd path, widening `valid_to` is only safe when the
        colliding members carry the SAME value set: then it is a genuine coverage
        union with no code over-claim. When the members carry DIVERGENT value sets
        the surviving (prefer-coded) set must keep ITS OWN member's `valid_to` —
        widening would extend that set's codes past the source row that defined
        them (Codex P2: `bu/SPEC` coded 1960–2014 + codeless 1960–2016 must NOT
        emit codes for 2015–2016, a code-search false positive). The dropped
        member's extra coverage is forfeited; for a research catalog an
        over-claimed code window (false positive) is worse than a dropped codeless
        tail.

          - SAME value set (both non-None equal, or both None) -> safe coverage
            union: widen `valid_to` to the WIDEST window. Keeping the first in
            delivery order would silently truncate coverage / close an
            open-ended state.
          - DIVERGENT value sets -> survivor keeps its own member window (no
            widening):
            * exactly one non-None (one member classified, the other was free
              text) -> the CODED member survives with its set + its window; a
              coalesce, not a conflict, so no warning. Prefer-coded makes the
              merge delivery-order-independent: a codeless member arriving first
              must not drop a later member's value set.
            * both non-None and DIFFERENT -> genuine conflict (e.g. LMED VARUTYP,
              MFR ICD): keep `prior` deterministically (delivery order) and WARN
              so the dropped alternative is auditable; never silent.

        The #401 `Värdemängd` fallback is what makes a same-state_id collision
        carry divergent value sets at all: the kodlista paths bucket a
        distinguishing code into a non-empty `value_set_version_label` (which
        changes the state_id), but each MERGED member carries its OWN inline
        `Värdemängd` and the fallback emits `value_set_version_label=None`, so
        two such members CAN collide on one state_id.

        The stored row is ALWAYS based on `prior`, with ONLY the survivor's
        `valid_to` and `value_set_id` reconciled onto it. Other scalars are NOT
        reconciled: `data_type` in particular is excluded from the state_id
        basis, and a warn-/allowlist-merge group can put members of different
        `data_type` under one variable_id — so two colliding members may differ
        in it. Basing the row on `prior` keeps `data_type` (and every other
        non-reconciled field) delivery-order-deterministic regardless of which
        side carried the value set; prefer-coded means a codeless `prior` adopts
        a later coded member's `value_set_id` (and that member's `valid_to`) but
        never its `data_type`.

        OVERLAP-SUPPRESSION POST-PASS (#401, complementary to `_collect`,
        VÄRDEMÄNGD-ONLY — gated to `kodlista is None`): `_collect` only reconciles
        members that collide on ONE state_id (same variant + same `valid_from`).
        Two members with DIFFERENT `valid_from` mint DIFFERENT state_ids, so
        `_collect` never compares them — yet they can still violate the build
        invariant ("one value_set per (variable, variant, period, column)",
        `validate.py`): OVERLAPPING windows on the same column with DISTINCT
        non-null value sets resolve a period to >1 value set. The kodlista paths
        avoid this by era-SEGMENTING on `Tidsperiod` (`_segment_windowed_codes` →
        non-overlapping segments; equal-content segments content-share one
        `value_set_id`) and the entity-registry path collapses to one state, so
        they own their own value sets and the post-pass NEVER touches them — a
        genuine kodlista conflict is caught by the build invariant / curation, not
        silently nulled here. Only the #401 Värdemängd binding has NO `Tidsperiod`
        to segment on, so two members whose own data windows (data_från/till) clip
        the shared kodlista to DIFFERENT code subsets over an overlap produce two
        distinct value sets there. After the member loop, the post-pass groups the
        buffered states by `(register_variant_id, delivery_column_name)` (mirroring
        the validator key; `variable_id` is fixed in this call), finds every
        overlapping pair with distinct non-null value sets, and nulls EVERY
        involved state's `value_set_id` back to code-less — the exact pre-#401
        behavior, so no regression — where the metadata is too ambiguous to
        auto-bind. One `sos_value_set_text_overlap` IRWarning per affected column
        makes the drop auditable. Disjoint-window multi-value-set variables
        (legitimate era changes) do NOT overlap and stay bound. Runs BEFORE the
        classification-candidate append so a suppressed state contributes its final
        (None) value_set_id.

        KODLISTA-WINS (#401 Fix B): a variable that HAS a kodlista sheet which was
        only skipped as unparseable (`raw_rows`) reaches the Värdemängd branch with
        `kodlista is None`, but `has_kodlista_sheet=True` keeps it code-less — never
        fabricate inline codes from Värdemängd when a (real) code list exists.
        Defensive: 0 corpus occurrences today, so inert on current output.
        """
        # Build the value-set rows for this variable's kodlista once.
        # entity_registry (MFR IVF_klinik): collapse to ONE state, per-code
        # valid_from/to. Otherwise: state-level era windowing per tidsperiod.
        seen_alias: set[tuple[int, str]] = set()
        # State buffer keyed by state_id; flushed after the member loop so the
        # reinsert sees one (widest-valid_to) row per state_id.
        states_by_id: dict[int, IRVariableState] = {}
        # #464 parallel buffer: the PENDING Värdemängd value-set identity per
        # surviving state_id, written only by the Värdemängd path (via `collect`'s
        # second arg). `None` = code-less (free text or suppressed); absent key =
        # an eager kodlista/entity-registry state (id already on the state row).
        # The survivor write (post-pass) reads this to materialize sets ONLY for
        # states that survive reconciliation — so dropped sets never orphan.
        pending_by_id: dict[int, PendingValueSet | None] = {}

        def _collect(
            obj: IRObject,
            pending: PendingValueSet | None | _NoPending = _NO_PENDING,
        ) -> Iterator[IRObject]:
            if isinstance(obj, IRVariableState):
                prior = states_by_id.get(obj.state_id)
                if prior is None:
                    states_by_id[obj.state_id] = obj
                    # #464: record the Värdemängd path's pending identity (the
                    # eager kodlista/entity-registry paths pass `_NO_PENDING` and
                    # carry their value_set_id on the state itself).
                    if not isinstance(pending, _NoPending):
                        pending_by_id[obj.state_id] = pending
                    return
                # kodlista paths (entity-registry + windowed): the #401
                # divergent-window / prefer-coded reconciliation is Värdemängd-ONLY
                # (`kodlista is None`). A kodlista-derived same-state_id collision
                # keeps the original pre-#401 behavior — ALWAYS widen `valid_to`,
                # keep `prior` (incl. its value_set_id). A distinguishing kodlista
                # code buckets into a non-None `value_set_version_label` that
                # changes the state_id, so a same-state_id kodlista collision
                # provably carries the SAME value set; widening is the only
                # reconciliation needed and never over-claims codes.
                if kodlista is not None:
                    states_by_id[obj.state_id] = prior.model_copy(
                        update={
                            "valid_to": _widest_valid_to(prior.valid_to, obj.valid_to)
                        }
                    )
                    return
                # #401 Värdemängd fallback (`kodlista is None`): two merged members
                # collided on one state_id, where value_set_version_label is always
                # None — so two such members CAN carry divergent inline Värdemängd
                # value sets. #464: reconcile on the BUFFERED `pending` identity
                # (member_hash) rather than `obj.value_set_id` (now always None —
                # the write is deferred). `prior_pending` is the survivor recorded
                # by an earlier `_collect`; `pending` is obj's identity.
                prior_pending = pending_by_id.get(obj.state_id)
                prior_hash = None if prior_pending is None else prior_pending[0]
                # In the `kodlista is None` branch every state came from the
                # Värdemängd path, which always passes a real pending (tuple or
                # None) — never `_NO_PENDING`. Normalize the sentinel to None so
                # `obj_pending` is the clean `PendingValueSet | None` the buffer
                # stores, and narrow the type for the survivor-buffer write below.
                obj_pending: PendingValueSet | None = (
                    None if isinstance(pending, _NoPending) else pending
                )
                obj_hash = None if obj_pending is None else obj_pending[0]
                # Same value set (equal member_hash, or both code-less): widening
                # `valid_to` is a safe coverage union — keeping the first in
                # delivery order would silently truncate coverage / close an
                # open-ended state, so reconcile to the WIDEST window. The pending
                # identity is unchanged (same content), so no rewrite needed.
                if prior_hash == obj_hash:
                    states_by_id[obj.state_id] = prior.model_copy(
                        update={
                            "valid_to": _widest_valid_to(prior.valid_to, obj.valid_to)
                        }
                    )
                    return
                # Divergent value sets: the surviving (prefer-coded) set keeps
                # ITS OWN member's window — never extend codes past their source
                # row (Codex P2). The dropped member's extra coverage is
                # forfeited; for a research catalog an over-claimed code window
                # (false positive) is worse than a dropped codeless tail. `prior`
                # wins ties and the both-coded conflict (delivery order); `obj`
                # wins only when `prior` is the codeless side. The survivor's
                # PENDING identity is buffered (so the deferred write materializes
                # only the kept set); the dropped side's identity is discarded and
                # never written -> no orphan.
                if prior_hash is None:
                    survivor_valid_to = obj.valid_to  # adopt obj's coded window
                    pending_by_id[obj.state_id] = obj_pending  # obj's coded set
                else:
                    survivor_valid_to = prior.valid_to  # keep prior's coded window
                    # prior's identity already buffered; leave it. (Both-coded
                    # conflict keeps prior; the obj identity is dropped, unwritten.)
                    if obj_hash is not None:
                        # both coded & DIFFERENT -> genuine conflict (kept first)
                        yield IRWarning(
                            entity_kind="variable",
                            entity_id=variable_id,
                            code="sos_value_set_text_conflict",
                            detail=(
                                f"{abbrev}/{members[0].name}: merged members share "
                                f"state_id {obj.state_id} but classify to different "
                                "Värdemängd value sets; kept first in delivery order"
                            ),
                        )
                # ALWAYS base the stored row on `prior`: `data_type` (and every
                # other non-reconciled scalar) is excluded from the state_id
                # basis, so a warn-/allowlist-merge group can put members of
                # different `data_type` under one variable_id — basing the row on
                # `obj` would silently flip it. Only `valid_to` is reconciled here
                # (value_set_id stays None until the deferred survivor write),
                # never `data_type`.
                states_by_id[obj.state_id] = prior.model_copy(
                    update={"valid_to": survivor_valid_to}
                )
                return
            yield obj

        for v in members:
            # Variant resolution: variant-less -> the synthesized _default;
            # else the member's deldatamängd token, routed through the curated
            # #211 token map when the token has no Deldatamängder-sheet row of
            # its own. A mapped token can name SEVERAL variants (LMED FDDD) —
            # the member emits into each.
            if variant_less:
                targets: list[tuple[int, SosDeldatamangd | None]] = [
                    (variant_id_of[None], None)
                ]
            else:
                names = DELDATAMANGD_TOKEN_MAP.get(
                    (abbrev, v.deldatamangd or ""), (v.deldatamangd,)
                )
                targets = []
                for target_name in names:
                    target_id = variant_id_of.get(target_name)
                    if target_id is None:
                        # P2#1: an UNCURATED variable-sheet deldatamängd token
                        # (absent from both the Deldatamängder sheet and
                        # DELDATAMANGD_TOKEN_MAP — e.g. a new workbook revision)
                        # can't be resolved to a variant. WARN so the drop is
                        # auditable, never silent; do NOT invent a mapping
                        # (e.g. by order) — skip the member after warning. The
                        # fix is always a curated token-map entry.
                        yield IRWarning(
                            entity_kind="variable",
                            entity_id=variable_id,
                            code="sos_deldatamangd_unresolved",
                            detail=(
                                f"{abbrev}/{v.name}: variable-sheet deldatamängd "
                                f"{v.deldatamangd} has no Deldatamängder-sheet "
                                "row; state dropped, add a DELDATAMANGD_TOKEN_MAP "
                                "entry (A4.4 curation)"
                            ),
                        )
                        continue
                    targets.append((target_id, deldat_by_name.get(target_name or "")))

            for variant_id, deldat in targets:
                yield from self._emit_member_states(
                    abbrev=abbrev,
                    variable_id=variable_id,
                    variant_id=variant_id,
                    deldat=deldat,
                    v=v,
                    seen_alias=seen_alias,
                    kodlista=kodlista,
                    entity_registry=entity_registry,
                    collect=_collect,
                    has_kodlista_sheet=has_kodlista_sheet,
                )

        # Overlap-suppression post-pass — Värdemängd-ONLY (`kodlista is None`),
        # complementary to `_collect`. `_collect` only reconciles members that
        # collide on ONE state_id (same variant + same valid_from). Two members
        # with DIFFERENT valid_from mint DIFFERENT state_ids, so they sit as
        # separate `states_by_id` entries and `_collect` never compares them — yet
        # OVERLAPPING windows on the same column carrying DISTINCT non-null value
        # sets still violate the build invariant (one value_set per (variable,
        # variant, period, column)). The #401 Värdemängd path binds one value set
        # over each member's WHOLE window with no Tidsperiod to segment on, so two
        # members whose windows overlap but whose Värdemängd cells classify to
        # different sets are unresolvable (e.g. bu/SPEC: a single-code 1960–2016
        # cell vs a full-enumeration 2001–2014 cell). Conservatively null EVERY
        # conflicting state back to code-less (the exact pre-#401 behavior — no
        # regression) and WARN once per column. Disjoint-window multi-value-set
        # variables (legitimate era changes) do NOT overlap and stay bound. Group
        # key mirrors the validator exactly; variable_id is fixed within this call.
        #
        # Gated to `kodlista is None`: the kodlista paths segment on Tidsperiod
        # (`_segment_windowed_codes`, non-overlapping segments) and the
        # entity-registry path collapses to one state, so they own their own value
        # sets and never need this suppression. NEVER null a kodlista-derived value
        # set here — a genuine kodlista conflict is caught by the build invariant /
        # curation, not silently dropped.
        if kodlista is None:
            by_column: dict[tuple[int, str | None], list[int]] = {}
            for sid, state in states_by_id.items():
                by_column.setdefault(
                    (state.register_variant_id, state.delivery_column_name), []
                ).append(sid)
            conflicted: set[int] = set()
            for sids in by_column.values():
                for i in range(len(sids)):
                    # #464: compare the PENDING member_hash (not value_set_id —
                    # the write is deferred, so all states carry None until the
                    # survivor write below). None pending = code-less, skip.
                    a_pending = pending_by_id.get(sids[i])
                    if a_pending is None:
                        continue
                    a = states_by_id[sids[i]]
                    a_hash = a_pending[0]
                    for j in range(i + 1, len(sids)):
                        b_pending = pending_by_id.get(sids[j])
                        if b_pending is None:
                            continue
                        b = states_by_id[sids[j]]
                        if a_hash == b_pending[0] or not _windows_overlap(
                            a.valid_from, a.valid_to, b.valid_from, b.valid_to
                        ):
                            # Equal hash = same set (not a conflict); disjoint
                            # windows = no invariant violation. Leave bound.
                            continue
                        conflicted.add(sids[i])
                        conflicted.add(sids[j])
            warned_columns: set[str | None] = set()
            for sid in conflicted:
                state = states_by_id[sid]
                col = state.delivery_column_name
                # #464: null the PENDING identity so the survivor write below
                # never materializes this set -> no orphan. The state stays
                # value_set_id=None (it already is, pre-survivor-write).
                pending_by_id[sid] = None
                if col not in warned_columns:
                    warned_columns.add(col)
                    yield IRWarning(
                        entity_kind="variable",
                        entity_id=variable_id,
                        code="sos_value_set_text_overlap",
                        detail=(
                            f"{abbrev}/{col}: overlapping members classified to "
                            "distinct Värdemängd value sets (no Tidsperiod to "
                            "segment on); binding dropped (code-less) where ambiguous"
                        ),
                    )

        # #464 SURVIVOR WRITE: now that reconciliation (`_collect` + the overlap
        # post-pass) has settled which Värdemängd states survive, materialize a
        # value set ONLY for each surviving state with a non-None pending identity.
        # `_ensure_value_set` is content-addressed (INSERT OR IGNORE on
        # UNIQUE(member_hash), cached in `_set_id_by_hash`), so two surviving
        # states with identical content still collapse to ONE shared value_set_id —
        # the deferred write hashed the SAME pairs, so the id is identical to the
        # old eager path. Dropped/nulled sets are simply never written -> the only
        # build-output change vs the eager path is that orphaned rows disappear.
        # Runs BEFORE the classification-candidate append so that append reads the
        # FINAL value_set_id (unchanged keying logic). Code-less survivors keep
        # value_set_id=None. (Eager kodlista/entity-registry states have no
        # pending_by_id entry and keep the value_set_id already on their row.)
        for sid, state in list(states_by_id.items()):
            pending = pending_by_id.get(sid)
            if pending is None:
                continue
            _member_hash, codes = pending
            value_set_id = self._ensure_value_set(codes)
            states_by_id[sid] = state.model_copy(update={"value_set_id": value_set_id})

        # Flush the reconciled states (one per state_id, widest valid_to).
        # For a RESOLVED variable, append one classification candidate per emitted
        # state keyed on (variable_id, state.value_set_id) — the backfill's state
        # key (value_set_id may be None for code-less ICD/ATC states; matched with
        # `IS`). valid_to reconciliation doesn't touch value_set_id, so the buffer
        # holds the final, deduped state set.
        if classification is not None and states_by_id:
            self._resolved_variables += 1
            for state in states_by_id.values():
                self.classification_candidates.append(
                    (variable_id, state.value_set_id, classification)
                )
                self._tagged_states += 1
        yield from states_by_id.values()

    def _emit_member_states(
        self,
        *,
        abbrev: str,
        variable_id: int,
        variant_id: int,
        deldat: SosDeldatamangd | None,
        v: SosVariable,
        seen_alias: set[tuple[int, str]],
        kodlista: SosKodlista | None,
        entity_registry: bool,
        collect: _CollectFn,
        has_kodlista_sheet: bool = False,
    ) -> Iterator[IRObject]:
        """One (member, resolved variant) emission: the alias row plus the
        member's era-windowed states routed through ``collect`` (the
        `_emit_states` same-state_id reconciliation buffer).

        #464 deferral: the Värdemängd branch does NOT write its value set here
        (eager writes orphan rows when reconciliation later drops the state's
        set). It computes the pending (member_hash, codes) — or `None` for a
        code-less state — and passes it to ``collect`` as the second arg, which
        reconciles same-state_id collisions and buffers the SURVIVOR's identity;
        `_emit_states` writes only surviving states' sets after reconciliation.
        kodlista + entity-registry paths are untouched: they still write eagerly
        (segmented/collapsed, so they never orphan) and call ``collect(obj)`` with
        no pending identity (`_NO_PENDING`).

        `has_kodlista_sheet`: the variable HAS a kodlista sheet that was skipped
        as unparseable (`raw_rows`), so `kodlista is None` here only because the
        sheet couldn't be parsed. Kodlista-wins: stay code-less rather than
        fabricating inline codes from Värdemängd (#401 Fix B)."""
        var_from = _iso_bound(v.data_from, end=False)
        var_to = _iso_bound(v.data_to, end=True)
        deldat_from = _iso_bound(deldat.data_from, end=False) if deldat else None
        deldat_to = _iso_bound(deldat.data_to, end=True) if deldat else None

        # variable_alias: the delivery column == the variable name for SOS.
        col = v.name
        if col and (variant_id, col) not in seen_alias:
            seen_alias.add((variant_id, col))
            yield IRVariableAlias(
                variable_id=variable_id,
                register_variant_id=variant_id,
                delivery_column_name=col,
            )

        var_bound = (var_from, var_to)
        deldat_bound = (deldat_from, deldat_to)
        # P2#2: the deldatamängd window is ADVISORY; the variable (+ code)
        # window is authoritative. `deldat_dropped` collects per-window
        # contradictions so we WARN once per (member, variant) — each variant
        # carries its own deldat window — not per code/state.
        deldat_dropped: list[bool] = []

        if kodlista is None:
            # No parsed `Kodlista_*` sheet for this variable. Fall back to the
            # inline `Värdemängd` cell (#401, the #373 deferral): when it is a
            # clean enumerated code list, promote it to a value set; otherwise it's
            # free text and the state stays code-less (exactly today's behavior).
            # Värdemängd carries no Tidsperiod, so there is no per-code windowing
            # and no era drift — one value set over the whole variable window.
            # EXCEPTION (#401 Fix B): if the variable HAS a kodlista sheet that was
            # only skipped as unparseable (`has_kodlista_sheet`), kodlista-wins —
            # never fabricate codes from Värdemängd; stay code-less.
            window, dropped = _intersect_advisory_deldat([var_bound], deldat_bound)
            if window is None:
                return
            if dropped:
                deldat_dropped.append(True)
            classified = (
                None
                if has_kodlista_sheet
                else _classify_value_set_text(v.value_set_text)
            )
            # #464 DEFERRAL: do NOT write the value set here. An eager
            # `_ensure_value_set` persists value_set/value_code/value_set_member
            # rows that `_emit_states` reconciliation can later orphan (a
            # divergent-collision drop in `_collect`, or an overlap-suppression
            # null in the post-pass) — orphaned codes then leak into unscoped
            # value search as context-less mapping_count=0 hits. Instead compute
            # the PENDING identity and hand it to `collect` (`_collect`), which
            # reconciles same-state_id collisions and records the SURVIVOR's
            # identity in `pending_by_id`; `_emit_states` then writes only
            # surviving states' sets. The state yields value_set_id=None;
            # `_emit_states` `model_copy`s the final id onto survivors.
            #
            # Footgun: the pending identity rides ALONGSIDE the state (not on it —
            # IRVariableState is a frozen content row), so it must travel through
            # `collect`, not via a `pending_by_id[state_id]` write here: two merged
            # members minting one state_id would otherwise clobber each other's
            # entry before `_collect` could compare them.
            pending: PendingValueSet | None
            if classified is None:
                pending = None  # free-text -> code-less
            else:
                # Värdemängd carries no Tidsperiod -> the whole-variable window.
                codes = [_value_code(code, label, window) for code, label in classified]
                # Hash the SAME pairs `_ensure_value_set` hashes (it builds
                # `[(c.code, c.label) for c in codes]`), so the deferred survivor
                # write content-shares onto the identical value_set_id.
                member_hash = _value_set_hash([(c.code, c.label) for c in codes])
                pending = (member_hash, codes)
            yield from collect(
                IRVariableState(
                    state_id=self._state_id(abbrev, variable_id, variant_id, window[0]),
                    variable_id=variable_id,
                    register_variant_id=variant_id,
                    valid_from=window[0] or _VALID_FROM_UNKNOWN,
                    valid_to=window[1],
                    data_type=_norm_data_type(v.data_type),
                    data_length=None,
                    delivery_column_name=col,
                    value_set_id=None,  # #464: assigned at survivor write
                    value_set_version_label=None,
                ),
                pending,
            )
        elif entity_registry:
            for obj in self._emit_entity_registry_state(
                abbrev=abbrev,
                variable_id=variable_id,
                variant_id=variant_id,
                v=v,
                col=col,
                kodlista=kodlista,
                var_bound=var_bound,
                deldat_bound=deldat_bound,
                deldat_dropped=deldat_dropped,
            ):
                yield from collect(obj)
        else:
            for obj in self._emit_windowed_states(
                abbrev=abbrev,
                variable_id=variable_id,
                variant_id=variant_id,
                v=v,
                col=col,
                kodlista=kodlista,
                var_bound=var_bound,
                deldat_bound=deldat_bound,
                deldat_dropped=deldat_dropped,
            ):
                yield from collect(obj)

        if deldat_dropped:
            yield IRWarning(
                entity_kind="variable",
                entity_id=variable_id,
                code="sos_deldatamangd_bound_contradicts_variable",
                detail=(
                    f"{abbrev}/{v.name}: deldatamängd window "
                    f"[{deldat_from or ''}..{deldat_to or ''}] excludes "
                    f"variable window [{var_from or ''}..{var_to or ''}]; "
                    "kept variable window"
                ),
            )

    def _emit_entity_registry_state(
        self,
        *,
        abbrev: str,
        variable_id: int,
        variant_id: int,
        v: SosVariable,
        col: str,
        kodlista: SosKodlista,
        var_bound: tuple[str | None, str | None],
        deldat_bound: tuple[str | None, str | None],
        deldat_dropped: list[bool],
    ) -> Iterator[IRObject]:
        """MFR IVF_klinik: ONE state over the full variable window, whose
        value_set is the UNION of every clinic code whose tidsperiod overlaps
        that window (the per-code window decides SURVIVAL — codes disjoint from
        the variable window drop — but is NOT persisted: value_code has no
        validity columns yet). The single state therefore exposes all surviving
        clinics across the whole window; per-code temporal precision is the
        deferred Path B refinement. Maintainer decision (2026-06-02): keep the
        collapse for A4.3b rather than fragmenting the entity registry into
        per-window states. See `_ensure_value_set`.

        P2#2: the variable bound + the code tidsperiod are authoritative; the
        deldat window is advisory and yields when it would empty the window.
        """
        codes: list[IRValueCode] = []
        for r in kodlista.rows:
            cf, ct = _parse_tidsperiod(r.tidsperiod)
            window, dropped = _intersect_advisory_deldat(
                [var_bound, (cf, ct)], deldat_bound
            )
            if window is None:
                continue  # code's window empty under var/code -> drop the code
            if dropped:
                deldat_dropped.append(True)
            # Per-code window (decides survival; not persisted on value_code).
            codes.append(_value_code(r.kod, r.beskrivning, window))
        if not codes:
            return
        value_set_id = self._ensure_value_set(codes)
        full, dropped = _intersect_advisory_deldat([var_bound], deldat_bound)
        if full is None:
            return
        if dropped:
            deldat_dropped.append(True)
        yield IRVariableState(
            state_id=self._state_id(abbrev, variable_id, variant_id, full[0]),
            variable_id=variable_id,
            register_variant_id=variant_id,
            valid_from=full[0] or _VALID_FROM_UNKNOWN,
            valid_to=full[1],
            data_type=_norm_data_type(v.data_type),
            data_length=None,
            delivery_column_name=col,
            value_set_id=value_set_id,
            value_set_version_label=None,
        )

    def _emit_windowed_states(
        self,
        *,
        abbrev: str,
        variable_id: int,
        variant_id: int,
        v: SosVariable,
        col: str,
        kodlista: SosKodlista,
        var_bound: tuple[str | None, str | None],
        deldat_bound: tuple[str | None, str | None],
        deldat_dropped: list[bool],
    ) -> Iterator[IRObject]:
        """Standard value-set DRIFT across per-code validity windows.

        SOS codes can carry HETEROGENEOUS, overlapping tidsperiod windows (a
        wide/open code coexisting with narrower sub-windows). Bucketing by EXACT
        window then produced OVERLAPPING value sets on one column — a period
        resolving to >1 value set (the co-delivery invariant — see DESIGN.md →
        Build-time triage (SCB)). Instead, sweep-line
        the 3-way-intersected windows into NON-OVERLAPPING period segments
        (`_segment_windowed_codes`), each carrying the union of codes live across
        it (the "Path B" per-period refinement), and emit one state per segment.

        P2#2: the variable bound + the code tidsperiod are authoritative; the
        deldat window is advisory and yields when it would empty the window.

        SCOPE: segmentation guarantees non-overlap WITHIN this call (one member).
        `_emit_states` calls it once per member of a merged variable, and two
        members can share a (variant, column); cross-member overlap is NOT
        structurally blocked here. It is empirically absent on today's corpus, but
        not impossible: members carry different `var_bound`/deldat clamps, so a
        period two members share can land in segments with DIFFERENT clamped
        `valid_from` (no exact `state_id` collision → `_emit_states` does not
        reconcile them). What keeps the invariant green in the common case is that equal
        code-content segments CONTENT-SHARE one `value_set_id`, and the invariant only fails
        on DISTINCT value sets. The residual hole: `_intersect_advisory_deldat` is
        a FALLBACK, not a pure clamp — when the deldat window would empty a code's
        window it is dropped and the wider authoritative window kept, so a code
        clipped out of member A (deldat honored) can survive in member B (deldat
        dropped); at a shared period on a shared column that is a DISTINCT value
        set. The **build invariant is the guarantee** that catches it (the
        build fails) — this segmentation is not a structural proof against it.
        """
        windowed: list[tuple[str | None, str | None, IRValueCode]] = []
        for r in kodlista.rows:
            cf, ct = _parse_tidsperiod(r.tidsperiod)
            window, dropped = _intersect_advisory_deldat(
                [var_bound, (cf, ct)], deldat_bound
            )
            if window is None:
                continue
            if dropped:
                deldat_dropped.append(True)
            # Per-code (3-way-intersected) segment window, carried into
            # `_segment_windowed_codes` to sweep into non-overlapping segments.
            windowed.append(
                (window[0], window[1], _value_code(r.kod, r.beskrivning, window))
            )
        segments = _segment_windowed_codes(windowed)
        if not segments:
            # No code survived windowing — emit a code-less state over the
            # variable window (deldat advisory) so the variable still has a state.
            window, dropped = _intersect_advisory_deldat([var_bound], deldat_bound)
            if window is None:
                return
            if dropped:
                deldat_dropped.append(True)
            yield IRVariableState(
                state_id=self._state_id(abbrev, variable_id, variant_id, window[0]),
                variable_id=variable_id,
                register_variant_id=variant_id,
                valid_from=window[0] or _VALID_FROM_UNKNOWN,
                valid_to=window[1],
                data_type=_norm_data_type(v.data_type),
                data_length=None,
                delivery_column_name=col,
                value_set_id=None,
                value_set_version_label=None,
            )
            return
        # Segments are non-overlapping, so each has a distinct valid_from — the
        # value_set_version_label (the segment range) only disambiguates a
        # same-valid_from collision ACROSS members of a merged variable; the
        # unique index keys on it.
        multi = len(segments) > 1
        for seg_from, seg_to, codes in segments:
            value_set_id = self._ensure_value_set(codes)
            label = "" if not multi else f"{seg_from or ''}/{seg_to or ''}"
            yield IRVariableState(
                state_id=self._state_id(
                    abbrev, variable_id, variant_id, seg_from, label
                ),
                variable_id=variable_id,
                register_variant_id=variant_id,
                valid_from=seg_from or _VALID_FROM_UNKNOWN,
                valid_to=seg_to,
                data_type=_norm_data_type(v.data_type),
                data_length=None,
                delivery_column_name=col,
                value_set_id=value_set_id,
                value_set_version_label=label or None,
            )

    @staticmethod
    def _state_id(
        abbrev: str,
        variable_id: int,
        variant_id: int,
        valid_from: str | None,
        version_label: str = "",
    ) -> int:
        """Mint a rebuild-stable state id from the unique-index 4-tuple basis
        (variable, variant, valid_from, version_label)."""
        return mint(
            "sos",
            "state",
            str(variable_id),
            str(variant_id),
            valid_from or "",
            version_label,
        )


def _first(group: list[SosVariable], attr: str) -> str | None:
    """First non-null value of ``attr`` across ``group`` (delivery order)."""
    for v in group:
        val = getattr(v, attr)
        if val:
            return val
    return None
