"""Tests for the Socialstyrelsen Excel parser.

Two tiers:

- Unit tests for helpers. Always run.
- Integration tests over the real input files under
  `reg_meta_build/input_data/Socialstyrelsen/`. Skipped when the directory
  is absent (CI, fresh checkouts) since the input is gitignored. These
  tests exist to catch regressions against real deliveries during local
  development.
"""

from __future__ import annotations

from pathlib import Path

import pytest

# The parser module imports lazily, so unit tests over pure helpers can run
# without openpyxl. Tests that actually load workbooks are gated by the
# `requires_openpyxl` mark below.
try:
    import openpyxl  # noqa: F401
except ImportError:
    HAS_OPENPYXL = False
else:
    HAS_OPENPYXL = True

requires_openpyxl = pytest.mark.skipif(
    not HAS_OPENPYXL, reason="openpyxl is required for this test"
)

from reg_meta_build.sources.sos import (  # noqa: E402
    SosDcatAp,
    SosParseError,
    _as_date,
    _as_int,
    _clean,
    _format_code,
    _normalise,
    parse_directory,
    parse_register_file,
)

# ---------------------------------------------------------------------------
# Unit tests
# ---------------------------------------------------------------------------


def test_normalise_strips_separators_and_case() -> None:
    assert _normalise("Metadata - Variabelnivå") == "metadatavariabelnivå"
    assert _normalise("Kodlista_DIAGNOS") == "kodlistadiagnos"
    assert _normalise("Metadata-Datamängd (DCAT-AP) ") == "metadatadatamängddcatap"


def test_clean_empty_returns_none() -> None:
    assert _clean(None) is None
    assert _clean("") is None
    assert _clean("   ") is None
    assert _clean("  hi ") == "hi"
    assert _clean(42) == "42"


def test_as_int_handles_common_shapes() -> None:
    assert _as_int(1964) == 1964
    assert _as_int("1964") == 1964
    assert _as_int("  1964 ") == 1964
    assert _as_int(1964.0) == 1964
    assert _as_int(1964.5) is None
    assert _as_int(None) is None
    assert _as_int("") is None
    assert _as_int("not a year") is None


def test_as_date_accepts_datetime_and_date() -> None:
    from datetime import date, datetime

    assert _as_date(datetime(2026, 3, 26)) == date(2026, 3, 26)
    assert _as_date(date(2026, 3, 26)) == date(2026, 3, 26)
    assert _as_date("2026-03-26") is None  # strings aren't implicitly parsed
    assert _as_date(None) is None


def test_lock_file_rejected() -> None:
    with pytest.raises(SosParseError):
        parse_register_file(Path("~$something.xlsx"))


def test_missing_file_raises() -> None:
    with pytest.raises(SosParseError):
        parse_register_file(Path("/does/not/exist.xlsx"))


def test_dcat_ap_extras_roundtrip() -> None:
    ap = SosDcatAp(title_sv="Foo", extras={"Unknown attribute": "value"})
    assert ap.extras == {"Unknown attribute": "value"}
    assert ap.title_sv == "Foo"


class _FakeCell:
    """Minimal duck-typed stand-in for an openpyxl cell — enough for
    `_format_code` to inspect `value` and `number_format` without pulling
    in the optional dep."""

    def __init__(self, value: object, number_format: str = "General") -> None:
        self.value = value
        self.number_format = number_format


def test_format_code_passes_strings_through() -> None:
    assert _format_code(_FakeCell("ABC123")) == "ABC123"
    assert _format_code(_FakeCell("  001 ")) == "001"
    assert _format_code(_FakeCell(None)) is None
    assert _format_code(_FakeCell("")) is None


def test_format_code_pads_int_with_pure_zero_format() -> None:
    # Excel stores e.g. "001" as int 1 with number_format "000"; without
    # consulting the format we'd lose the leading zeros and corrupt code
    # identity. Only pure-zero formats are treated as code padding.
    assert _format_code(_FakeCell(1, "000")) == "001"
    assert _format_code(_FakeCell(12, "000")) == "012"
    assert _format_code(_FakeCell(123, "000")) == "123"
    assert _format_code(_FakeCell(7, "General")) == "7"
    assert _format_code(_FakeCell(7.0, "00")) == "07"
    assert _format_code(_FakeCell(7.5, "General")) == "7.5"


def _write_minimal_workbook(
    path: Path,
    *,
    kod_rows: list[tuple[str | None, object, str | None]] | None = None,
    kod_format: str = "@",
    var_header: list[str] | None = None,
) -> None:
    """Write a minimal SoS-shaped workbook at `path`. Optionally include a
    Kodlista_TEST sheet with given rows and a number_format on the Kod
    column. `var_header` overrides the default `["Variabelnamn"]` header
    on the variable sheet (used by tests that exercise malformed-shape
    failure paths)."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Generell information"
    wb.create_sheet("Metadata-Datamängd (DCAT-AP)")
    var_ws = wb.create_sheet("Metadata - Variabelnivå")
    var_ws.append(var_header if var_header is not None else ["Variabelnamn"])
    var_ws.append(["TESTVAR"])
    if kod_rows is not None:
        kod_ws = wb.create_sheet("Kodlista_TEST")
        kod_ws.append(["Tidsperiod", "Kod", "Beskrivning"])
        for tp, kod, desc in kod_rows:
            kod_ws.append([tp, kod, desc])
        for row in kod_ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = kod_format
    wb.save(path)


@requires_openpyxl
def test_zero_padded_kod_round_trips_through_workbook(tmp_path: Path) -> None:
    p = tmp_path / "Test.xlsx"
    _write_minimal_workbook(
        p,
        kod_rows=[("2024", 1, "first"), ("2024", 12, "twelfth")],
        kod_format="000",
    )
    result = parse_register_file(p)
    assert len(result.kodlistor) == 1
    assert [r.kod for r in result.kodlistor[0].rows] == ["001", "012"]


@requires_openpyxl
def test_uppercase_xlsx_extension_picked_up_by_directory_parse(
    tmp_path: Path,
) -> None:
    p = tmp_path / "TEST.XLSX"
    _write_minimal_workbook(p)
    results = parse_directory(tmp_path)
    assert len(results) == 1
    assert results[0].source_file.name == "TEST.XLSX"


@requires_openpyxl
def test_directory_passed_as_file_rejected(tmp_path: Path) -> None:
    with pytest.raises(SosParseError, match="not a regular file"):
        parse_register_file(tmp_path)


@requires_openpyxl
def test_variable_sheet_without_variabelnamn_header_raises(tmp_path: Path) -> None:
    # An upstream rename or malformed delivery would leave the varsheet
    # without a Variabelnamn column. Silently returning zero variables
    # would hide the problem, so we fail fast here.
    p = tmp_path / "Test.xlsx"
    _write_minimal_workbook(p, var_header=["Foo", "Bar"])
    with pytest.raises(SosParseError, match="Variabelnamn"):
        parse_register_file(p)


@requires_openpyxl
def test_unsupported_xls_format_wrapped_as_sos_parse_error(tmp_path: Path) -> None:
    # openpyxl raises InvalidFileException for `.xls`/`.xlsb`; surface as
    # SosParseError so the parser's contract holds for common wrong inputs.
    p = tmp_path / "test.xls"
    p.write_bytes(b"")
    with pytest.raises(SosParseError, match="does not support"):
        parse_register_file(p)


# ---------------------------------------------------------------------------
# Synthetic-workbook parser coverage
#
# These replace the deleted real-delivery integration tests: they pin the same
# parser branches against a hand-built workbook instead of the gitignored real
# deliveries — DCAT-AP field map, full variable fields (incl. Länk kodverk),
# Deldatamängder parsing, a per-row Variabelnamn kodlista (MFR shape), an
# unrecognised kodlista (-> raw_rows), Generell key/values, phantom rows, and
# directory lock-file skipping. Real-corpus drift is deliberately NOT a CI
# concern; a maintainer's actual `build-db` over the real input surfaces that.
# ---------------------------------------------------------------------------


def _write_rich_workbook(path: Path) -> None:
    """Write a multi-sheet SoS-shaped workbook exercising the parser's branches.

    Shapes mirror the live deliveries the deleted integration tests pinned: a
    Generell key/value sheet, a full DCAT-AP attribute map, a Deldatamängder
    sheet, a variable sheet with the full column set, a Kodlista with a per-row
    ``Variabelnamn`` column (MFR shape), an unrecognised Kodlista (no
    Tidsperiod/Kod header -> raw_rows), and trailing phantom rows.
    """
    import openpyxl

    wb = openpyxl.Workbook()

    gen = wb.active
    gen.title = "Generell information"
    gen.append(["", "Om metadatamallen", None])  # section header (value None)
    gen.append(["", "Version", "2.1"])
    gen.append(["", "Om datamängden version", None])  # section header
    gen.append(["", "Datamängd", "Patientregistret-syntet"])
    gen.append(["", "Version", "2024:2"])
    gen.append(["", "E-post", "kontakt@example.se"])

    dcat = wb.create_sheet("Metadata-Datamängd (DCAT-AP)")
    dcat.append(["Attribut SoS-metadata", "Definition", "Svenska", "Engelska"])
    dcat.append(["Titel", "", "Patientregistret-syntet", "Synthetic Patient Register"])
    dcat.append(
        ["Beskrivning", "", "En syntetisk beskrivning", "A synthetic description"]
    )
    dcat.append(["Tidsperiod", "", "1964-", ""])
    dcat.append(["Utgivare", "", "Socialstyrelsen", ""])
    dcat.append(["Ingångssida", "", "https://example.se/par", ""])
    dcat.append(["Åtkomsträttigheter", "", "Begränsad åtkomst", ""])
    dcat.append(["Tillämplig lagstiftning", "", "Lag om hälsodataregister", ""])
    dcat.append(["Okänt attribut", "", "extra-värde", ""])  # -> dcat_ap.extras

    dd = wb.create_sheet("Deldatamängder och datavyer")
    dd.append(["Deldatamängdsnamn", "Deldatamängdsetikett", "Data från", "Data till"])
    dd.append(["PAR_OV", "Öppenvård", 2001, 2020])
    dd.append(["PAR_SV", "Slutenvård", 1987, 2020])

    var = wb.create_sheet("Metadata - Variabelnivå")
    var.append(
        [
            "Deldatamängdsnamn",
            "Variabelnamn",
            "Variabeletikett",
            "Variabelbeskrivning",
            "Objekttyp",
            "Värdemängd",
            "Länk kodverk",
            "Datatyp",
            "Data från",
            "Data till",
        ]
    )
    var.append(
        [
            "PAR_OV",
            "HDIA",
            "Huvuddiagnos",
            "Huvuddiagnoskod",
            "Vårdkontakt",
            "Se kodlista",
            "ICD-10",
            "Sträng (text)",
            2001,
            2020,
        ]
    )
    var.append(
        [
            "PAR_OV",
            "KON",
            "Kön",
            "Personens kön",
            "Person",
            "",
            "",
            "Heltal",
            2001,
            2020,
        ]
    )
    # Trailing phantom/empty rows must not become variables.
    var.append([None] * 10)
    var.append([None] * 10)

    # MFR shape: a per-row Variabelnamn column alongside Tidsperiod/Kod.
    kod = wb.create_sheet("Kodlista_BUTSATT")
    kod.append(["Tidsperiod", "Kod", "Beskrivning", "Variabelnamn"])
    kod.append(["2001-2020", "1", "Ja", "BUTSATT"])
    kod.append(["2001-2020", "0", "Nej", "BUTSATT"])

    # Unrecognised shape: no Tidsperiod/Kod header -> structured rows skipped,
    # raw content preserved.
    weird = wb.create_sheet("Kodlista_WEIRD")
    weird.append(["Sjukhuskatalog", "", ""])
    weird.append(["Karolinska", "Stockholm", "01"])
    weird.append(["Sahlgrenska", "Göteborg", "02"])

    wb.save(path)


@requires_openpyxl
def test_synthetic_generell_fields(tmp_path: Path) -> None:
    p = tmp_path / "Metadata Syntet (PAR)_webb.xlsx"
    _write_rich_workbook(p)
    reg = parse_register_file(p)
    assert reg.template_version == "2.1"
    assert reg.dataset_name == "Patientregistret-syntet"
    assert reg.dataset_version == "2024:2"
    assert reg.contact_email == "kontakt@example.se"


@requires_openpyxl
def test_synthetic_dcat_ap_field_map(tmp_path: Path) -> None:
    p = tmp_path / "Metadata Syntet (PAR)_webb.xlsx"
    _write_rich_workbook(p)
    d: SosDcatAp = parse_register_file(p).dcat_ap
    assert d.title_sv == "Patientregistret-syntet"
    assert d.title_en == "Synthetic Patient Register"
    assert d.description_sv and d.description_en
    assert d.temporal_coverage_sv == "1964-"
    assert d.publisher_sv == "Socialstyrelsen"
    assert d.landing_page_sv and d.landing_page_sv.startswith("https://")
    assert d.access_rights_sv
    assert d.legislation_sv and "hälsodataregister" in d.legislation_sv.lower()
    assert d.extras.get("Okänt attribut") == "extra-värde"


@requires_openpyxl
def test_synthetic_deldatamangder_parsed(tmp_path: Path) -> None:
    p = tmp_path / "Metadata Syntet (PAR)_webb.xlsx"
    _write_rich_workbook(p)
    by_name = {d.name: d for d in parse_register_file(p).deldatamangder}
    assert set(by_name) == {"PAR_OV", "PAR_SV"}
    assert by_name["PAR_OV"].label == "Öppenvård"
    assert by_name["PAR_OV"].data_from == 2001
    assert by_name["PAR_OV"].data_to == 2020


@requires_openpyxl
def test_synthetic_variable_fields_and_external_classification(tmp_path: Path) -> None:
    p = tmp_path / "Metadata Syntet (PAR)_webb.xlsx"
    _write_rich_workbook(p)
    hdia = next(v for v in parse_register_file(p).variables if v.name == "HDIA")
    assert hdia.deldatamangd == "PAR_OV"
    assert hdia.label == "Huvuddiagnos"
    assert hdia.data_type == "Sträng (text)"
    assert hdia.object_type == "Vårdkontakt"
    assert hdia.external_classification == "ICD-10"
    assert hdia.data_from == 2001
    assert hdia.data_to == 2020


@requires_openpyxl
def test_synthetic_kodlista_per_row_variable_column(tmp_path: Path) -> None:
    # MFR shape: the Kodlista carries a per-row Variabelnamn column.
    p = tmp_path / "Metadata Syntet (PAR)_webb.xlsx"
    _write_rich_workbook(p)
    reg = parse_register_file(p)
    butsatt = next(k for k in reg.kodlistor if k.variable_hint.lower() == "butsatt")
    assert butsatt.rows
    assert butsatt.rows[0].variable_name == "BUTSATT"
    assert {r.kod for r in butsatt.rows} == {"1", "0"}


@requires_openpyxl
def test_synthetic_unrecognised_kodlista_preserves_raw_rows(tmp_path: Path) -> None:
    p = tmp_path / "Metadata Syntet (PAR)_webb.xlsx"
    _write_rich_workbook(p)
    reg = parse_register_file(p)
    weird = next(k for k in reg.kodlistor if k.variable_hint.lower() == "weird")
    assert weird.rows == ()
    assert weird.raw_rows
    assert any("Karolinska" in str(c) for row in weird.raw_rows for c in row)


@requires_openpyxl
def test_synthetic_phantom_rows_do_not_inflate_variable_count(tmp_path: Path) -> None:
    # openpyxl's max_row is unreliable; `_row_iter` skips empty rows so trailing
    # phantom rows never become variables.
    p = tmp_path / "Metadata Syntet (PAR)_webb.xlsx"
    _write_rich_workbook(p)
    assert {v.name for v in parse_register_file(p).variables} == {"HDIA", "KON"}


@requires_openpyxl
def test_synthetic_register_without_deldatamangder_warns(tmp_path: Path) -> None:
    # A workbook with no Deldatamängder sheet parses with an implicit-subset
    # warning (the LSS/BU/SOL shape the adapter turns into a _default variant).
    p = tmp_path / "Metadata Minimal (XXX)_webb.xlsx"
    _write_minimal_workbook(p)
    reg = parse_register_file(p)
    assert reg.deldatamangder == ()
    assert any("Deldatamängder" in w for w in reg.warnings)


@requires_openpyxl
def test_synthetic_parse_directory_skips_lock_files(tmp_path: Path) -> None:
    _write_rich_workbook(tmp_path / "Metadata Syntet (PAR)_webb.xlsx")
    (tmp_path / "~$Metadata Syntet (PAR)_webb.xlsx").write_bytes(b"lock")
    results = parse_directory(tmp_path)
    assert len(results) == 1
    assert not results[0].source_file.name.startswith("~$")
