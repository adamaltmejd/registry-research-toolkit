"""Synthetic Socialstyrelsen workbook fixtures for CI build coverage.

The SCB side has `_csv_fixtures.write_scb_input`; this is its SOS analog. It
materializes small `.xlsx` register workbooks under ``<input_dir>/Socialstyrelsen/``
so the SOS adapter (`reg_meta_build.sources.sos`) and a combined ``scb,sos`` build
run end-to-end in CI WITHOUT the gitignored 14GB real deliveries.

The workbooks are shaped to satisfy `sos.parse_register_file`: sheet names are
matched on normalised tokens (`_find_sheet`), so casing/whitespace is flexible.
The default fixture set exercises the two structural shapes the adapter branches
on:

  - ``(SYN)`` — a register WITH a Deldatamängder sheet (real variants) and a
    Kodlista sheet (a value set). ``DIAGNOS`` appears under both deldatamängder,
    so it MERGES to one variable with one state per variant.
  - ``(SYT)`` — a register WITHOUT a Deldatamängder sheet, which the adapter
    detects and collapses to a synthesized ``_default`` variant.

The abbrev the adapter mints from is the parenthesized code in the FILENAME stem
(`_sos_abbrev`), so the file names carry ``(SYN)`` / ``(SYT)``.
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import TYPE_CHECKING

if TYPE_CHECKING:
    import sqlite3
    from pathlib import Path


@dataclass(frozen=True)
class _Deldat:
    name: str
    label: str | None = None
    description: str | None = None
    data_from: int | None = None
    data_to: int | None = None
    # Aggregeringsnivå controlled-vocab cell; 'Ej relevant' + a 'Styrtabell …'
    # label flags a styrtabell decode table the adapter excludes (#373).
    aggregation_level: str | None = None


@dataclass(frozen=True)
class _Var:
    name: str
    deldatamangd: str | None = None
    label: str | None = None
    description: str | None = None
    data_type: str = "Sträng (text)"
    data_from: int | None = None
    data_to: int | None = None
    # Raw `Länk kodverk` free-text — the signal the classification resolver
    # parses (SosVariable.external_classification). None emits a blank cell.
    external_classification: str | None = None


@dataclass(frozen=True)
class _Kodlista:
    """A Kodlista_<variable_hint> sheet. ``rows`` are (tidsperiod, kod, desc)."""

    variable_hint: str
    rows: tuple[tuple[str | None, str, str | None], ...]


@dataclass(frozen=True)
class _Register:
    abbrev: str  # parenthesized filename code, e.g. "SYN"
    title_sv: str
    description_sv: str | None
    variables: tuple[_Var, ...]
    deldatamangder: tuple[_Deldat, ...] = ()
    kodlistor: tuple[_Kodlista, ...] = ()
    dataset_version: str | None = "2024:1"


# Default fixture set. Two registers spanning the adapter's structural branches.
DEFAULT_REGISTERS: tuple[_Register, ...] = (
    _Register(
        abbrev="SYN",
        title_sv="Syntetiskt register",
        description_sv="Ett syntetiskt SOS-register för testbygget.",
        deldatamangder=(
            _Deldat(
                "SYN_A",
                label="Vy A",
                description="Första vyn",
                data_from=2005,
                data_to=2015,
            ),
            _Deldat("SYN_B", label="Vy B", description="Andra vyn", data_from=2010),
        ),
        variables=(
            _Var(
                "DIAGNOS",
                deldatamangd="SYN_A",
                label="Diagnoskod",
                description="ICD-kod",
                data_type="Sträng (text)",
                data_from=2005,
                data_to=2015,
            ),
            _Var(
                "DIAGNOS",
                deldatamangd="SYN_B",
                label="Diagnoskod",
                description="ICD-kod",
                data_type="Sträng (text)",
                data_from=2010,
            ),
            _Var(
                "KON",
                deldatamangd="SYN_A",
                label="Kön",
                description="Personens kön",
                data_type="Heltal",
                data_from=2005,
                data_to=2015,
            ),
        ),
        kodlistor=(
            _Kodlista(
                "DIAGNOS",
                rows=(
                    ("2005-2015", "A01", "Diagnos A"),
                    ("2005-2015", "B02", "Diagnos B"),
                ),
            ),
        ),
    ),
    _Register(
        abbrev="SYT",
        title_sv="Syntetiskt variantlöst register",
        description_sv="Saknar Deldatamängder-blad; adaptern syntetiserar _default.",
        # No deldatamangder sheet -> variant-less -> _default synthesis.
        variables=(
            _Var(
                "LOPNR",
                label="Löpnummer",
                description="Radens löpnummer",
                data_type="Heltal",
                data_from=2000,
                data_to=2020,
            ),
        ),
    ),
)


# A register that triggers the ("par", "ATC") KNOWN_SPLIT_ALLOWLIST split: ATC
# arrives under two deldatamängder with incompatible data_types, so the adapter
# splits it into two sibling variables and records a related-to edge between
# them. With slugs populated the build materializes that edge into
# `variable_related_to` — the surface the P3#1 leaked-loop-var regression
# polluted. Append to DEFAULT_REGISTERS for the with-slugs combined test.
PAR_SPLIT_REGISTER = _Register(
    abbrev="PAR",
    title_sv="Syntetiskt patientregister",
    description_sv="Split-test: ATC under två deldatamängder med olika datatyp.",
    deldatamangder=(
        _Deldat("PAR_OV", label="Öppenvård", data_from=2001),
        _Deldat("PAR_SV", label="Slutenvård", data_from=2001),
    ),
    variables=(
        _Var("ATC", deldatamangd="PAR_OV", data_type="Sträng (text)", data_from=2001),
        _Var("ATC", deldatamangd="PAR_SV", data_type="Heltal", data_from=2001),
    ),
)


def _write_register(path: Path, reg: _Register) -> None:
    import openpyxl

    wb = openpyxl.Workbook()

    # -- Generell information: col B label, col C value (see _parse_generell).
    gen = wb.active
    gen.title = "Generell information"
    gen.append(["", "Om metadatamallen", None])  # section header (value None)
    gen.append(["", "Version", "1.0"])
    gen.append(["", "Om datamängden version", None])  # section header
    gen.append(["", "Datamängd", reg.title_sv])
    gen.append(["", "Version", reg.dataset_version])
    gen.append(["", "E-post", "syntetisk@example.se"])

    # -- DCAT-AP: header row then (attr, _, svenska, engelska).
    dcat = wb.create_sheet("Metadata-Datamängd (DCAT-AP)")
    dcat.append(["Attribut", "Beskrivning", "Svenska", "Engelska"])
    dcat.append(["Titel", None, reg.title_sv, None])
    if reg.description_sv is not None:
        dcat.append(["Beskrivning", None, reg.description_sv, None])

    # -- Deldatamängder (optional): absence triggers _default synthesis.
    if reg.deldatamangder:
        dd = wb.create_sheet("Deldatamängder och datavyer")
        dd.append(
            [
                "Deldatamängdsnamn",
                "Deldatamängdsetikett",
                "Deldatamängdsbeskrivning",
                "Data från",
                "Data till",
                "Aggregeringsnivå",
            ]
        )
        for d in reg.deldatamangder:
            dd.append(
                [
                    d.name,
                    d.label,
                    d.description,
                    d.data_from,
                    d.data_to,
                    d.aggregation_level,
                ]
            )

    # -- Metadata - Variabelnivå (required).
    var = wb.create_sheet("Metadata - Variabelnivå")
    var.append(
        [
            "Deldatamängdsnamn",
            "Variabelnamn",
            "Variabeletikett",
            "Variabelbeskrivning",
            "Datatyp",
            "Länk kodverk",
            "Data från",
            "Data till",
        ]
    )
    for v in reg.variables:
        var.append(
            [
                v.deldatamangd,
                v.name,
                v.label,
                v.description,
                v.data_type,
                v.external_classification,
                v.data_from,
                v.data_to,
            ]
        )

    # -- Kodlista_<hint> (optional): per-variable value sets.
    for k in reg.kodlistor:
        ws = wb.create_sheet(f"Kodlista_{k.variable_hint}")
        ws.append(["Tidsperiod", "Kod", "Beskrivning"])
        for tp, kod, desc in k.rows:
            ws.append([tp, kod, desc])

    wb.save(path)


def write_sos_input(
    input_dir: Path, *, registers: tuple[_Register, ...] = DEFAULT_REGISTERS
) -> Path:
    """Materialize the synthetic SOS workbook set under ``<input_dir>/Socialstyrelsen/``.

    Returns the Socialstyrelsen subdirectory path. ``registers`` defaults to
    ``DEFAULT_REGISTERS``; pass a custom tuple to build alternate shapes.
    """
    sos_dir = input_dir / "Socialstyrelsen"
    sos_dir.mkdir(parents=True, exist_ok=True)
    for reg in registers:
        # Filename stem carries the parenthesized abbrev `_sos_abbrev` reads.
        path = sos_dir / f"Metadata {reg.title_sv} ({reg.abbrev})_webb.xlsx"
        _write_register(path, reg)
    return sos_dir


# ---------------------------------------------------------------------------
# Synthetic slug curation
#
# `build_db(skip_slugs=False)` calls `populate_slugs(strict=True)`, which refuses
# unless EVERY live register + register_variant carries a curated slug entry
# (register/variant slugs don't auto-derive; only variable slugs do). Hand-keying
# the minted SOS ids would be brittle, so we harvest them from a throwaway no-slug
# "probe" build (ids are deterministic, so the probe and the real with-slugs build
# share them) and emit one entry per row. The TOML keys are the source ids, which
# equal the register_id / register_variant_id for BOTH providers (SCB uses its
# low-band source ids directly; SOS's minted id IS its source id — see sos.toml).
# ---------------------------------------------------------------------------


def _kebab(name: str) -> str:
    """Best-effort valid slug (`^[a-z][a-z0-9-]*[a-z0-9]$`) from a display name.

    Legibility is irrelevant for tests; this only has to satisfy `validate_slug`.
    """
    import re

    s = re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")
    if not s:
        return "x"
    if not s[0].isalpha():
        s = "r" + s
    if len(s) > 1 and not s[-1].isalnum():
        s = s + "x"
    return s


def write_slug_dir_from_db(conn: sqlite3.Connection, slug_dir: Path) -> Path:
    """Generate ``<provider>.toml`` slug files covering every register + variant
    in ``conn`` (a built reg_meta DB), so a with-slugs build over the same inputs
    passes `populate_slugs(strict=True)`. Returns ``slug_dir``.
    """
    slug_dir.mkdir(parents=True, exist_ok=True)

    registers = conn.execute(
        "SELECT r.register_id, p.slug, r.name FROM register r "
        "JOIN provider p ON r.provider_id = p.provider_id ORDER BY r.register_id"
    ).fetchall()
    variants = conn.execute(
        "SELECT rv.register_id, rv.register_variant_id, rv.name, p.slug "
        "FROM register_variant rv JOIN register r USING (register_id) "
        "JOIN provider p ON r.provider_id = p.provider_id "
        "ORDER BY rv.register_id, rv.register_variant_id"
    ).fetchall()

    variants_by_reg: dict[int, list[tuple[int, str | None]]] = {}
    for register_id, variant_id, vname, _prov in variants:
        variants_by_reg.setdefault(register_id, []).append((variant_id, vname))

    lines_by_provider: dict[str, list[str]] = {}
    reg_slugs_seen: dict[str, set[str]] = {}
    for register_id, provider, rname in registers:
        lines = lines_by_provider.setdefault(provider, [])
        seen = reg_slugs_seen.setdefault(provider, set())
        rslug = _uniquify(_kebab(rname or f"reg{register_id}"), seen)
        lines.append(f'[register."{register_id}"]')
        lines.append(f'slug = "{rslug}"')
        lines.append("")
        variant_seen: set[str] = set()
        for variant_id, vname in variants_by_reg.get(register_id, []):
            # The synthesized default variant keeps the reserved `_default` slug.
            vslug = (
                "_default"
                if vname == "_default"
                else _uniquify(_kebab(vname or f"v{variant_id}"), variant_seen)
            )
            lines.append(f'[register_variant."{register_id}.{variant_id}"]')
            lines.append(f'slug = "{vslug}"')
            lines.append("")

    for provider, lines in lines_by_provider.items():
        (slug_dir / f"{provider}.toml").write_text(
            "\n".join(lines) + "\n", encoding="utf-8"
        )
    return slug_dir


def _uniquify(base: str, seen: set[str]) -> str:
    slug = base
    n = 2
    while slug in seen:
        slug = f"{base}-{n}"
        n += 1
    seen.add(slug)
    return slug
