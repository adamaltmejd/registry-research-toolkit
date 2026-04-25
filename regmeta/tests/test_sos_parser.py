"""Tests for the Socialstyrelsen Excel parser.

Two tiers:

- Unit tests for helpers. Always run.
- Integration tests over the real input files under
  `regmeta/input_data/Socialstyrelsen/`. Skipped when the directory is
  absent (CI, fresh checkouts) since the input is gitignored. These
  tests exist to catch regressions against real deliveries during local
  development.
"""

from __future__ import annotations

import os
from pathlib import Path

import pytest

# The parser module imports lazily, so unit tests over pure helpers can run
# without openpyxl. Tests that actually load workbooks are gated by the
# `requires_openpyxl` mark below.
try:
    import openpyxl  # noqa: F401
except ImportError:
    HAS_OPENPYXL = False
else:
    HAS_OPENPYXL = True

requires_openpyxl = pytest.mark.skipif(
    not HAS_OPENPYXL, reason="openpyxl is required for this test"
)

from regmeta.sources.sos import (  # noqa: E402
    SosDcatAp,
    SosParseError,
    SosRegister,
    _as_date,
    _as_int,
    _clean,
    _format_code,
    _normalise,
    parse_directory,
    parse_register_file,
)

# ---------------------------------------------------------------------------
# Unit tests
# ---------------------------------------------------------------------------


def test_normalise_strips_separators_and_case() -> None:
    assert _normalise("Metadata - Variabelnivå") == "metadatavariabelnivå"
    assert _normalise("Kodlista_DIAGNOS") == "kodlistadiagnos"
    assert _normalise("Metadata-Datamängd (DCAT-AP) ") == "metadatadatamängddcatap"


def test_clean_empty_returns_none() -> None:
    assert _clean(None) is None
    assert _clean("") is None
    assert _clean("   ") is None
    assert _clean("  hi ") == "hi"
    assert _clean(42) == "42"


def test_as_int_handles_common_shapes() -> None:
    assert _as_int(1964) == 1964
    assert _as_int("1964") == 1964
    assert _as_int("  1964 ") == 1964
    assert _as_int(1964.0) == 1964
    assert _as_int(1964.5) is None
    assert _as_int(None) is None
    assert _as_int("") is None
    assert _as_int("not a year") is None


def test_as_date_accepts_datetime_and_date() -> None:
    from datetime import date, datetime

    assert _as_date(datetime(2026, 3, 26)) == date(2026, 3, 26)
    assert _as_date(date(2026, 3, 26)) == date(2026, 3, 26)
    assert _as_date("2026-03-26") is None  # strings aren't implicitly parsed
    assert _as_date(None) is None


def test_lock_file_rejected() -> None:
    with pytest.raises(SosParseError):
        parse_register_file(Path("~$something.xlsx"))


def test_missing_file_raises() -> None:
    with pytest.raises(SosParseError):
        parse_register_file(Path("/does/not/exist.xlsx"))


def test_dcat_ap_extras_roundtrip() -> None:
    ap = SosDcatAp(title_sv="Foo", extras={"Unknown attribute": "value"})
    assert ap.extras == {"Unknown attribute": "value"}
    assert ap.title_sv == "Foo"


class _FakeCell:
    """Minimal duck-typed stand-in for an openpyxl cell — enough for
    `_format_code` to inspect `value` and `number_format` without pulling
    in the optional dep."""

    def __init__(self, value: object, number_format: str = "General") -> None:
        self.value = value
        self.number_format = number_format


def test_format_code_passes_strings_through() -> None:
    assert _format_code(_FakeCell("ABC123")) == "ABC123"
    assert _format_code(_FakeCell("  001 ")) == "001"
    assert _format_code(_FakeCell(None)) is None
    assert _format_code(_FakeCell("")) is None


def test_format_code_pads_int_with_pure_zero_format() -> None:
    # Excel stores e.g. "001" as int 1 with number_format "000"; without
    # consulting the format we'd lose the leading zeros and corrupt code
    # identity. Only pure-zero formats are treated as code padding.
    assert _format_code(_FakeCell(1, "000")) == "001"
    assert _format_code(_FakeCell(12, "000")) == "012"
    assert _format_code(_FakeCell(123, "000")) == "123"
    assert _format_code(_FakeCell(7, "General")) == "7"
    assert _format_code(_FakeCell(7.0, "00")) == "07"
    assert _format_code(_FakeCell(7.5, "General")) == "7.5"


def _write_minimal_workbook(
    path: Path,
    *,
    kod_rows: list[tuple[str | None, object, str | None]] | None = None,
    kod_format: str = "@",
    var_header: list[str] | None = None,
) -> None:
    """Write a minimal SoS-shaped workbook at `path`. Optionally include a
    Kodlista_TEST sheet with given rows and a number_format on the Kod
    column. `var_header` overrides the default `["Variabelnamn"]` header
    on the variable sheet (used by tests that exercise malformed-shape
    failure paths)."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "Generell information"
    wb.create_sheet("Metadata-Datamängd (DCAT-AP)")
    var_ws = wb.create_sheet("Metadata - Variabelnivå")
    var_ws.append(var_header if var_header is not None else ["Variabelnamn"])
    var_ws.append(["TESTVAR"])
    if kod_rows is not None:
        kod_ws = wb.create_sheet("Kodlista_TEST")
        kod_ws.append(["Tidsperiod", "Kod", "Beskrivning"])
        for tp, kod, desc in kod_rows:
            kod_ws.append([tp, kod, desc])
        for row in kod_ws.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                cell.number_format = kod_format
    wb.save(path)


@requires_openpyxl
def test_zero_padded_kod_round_trips_through_workbook(tmp_path: Path) -> None:
    p = tmp_path / "Test.xlsx"
    _write_minimal_workbook(
        p,
        kod_rows=[("2024", 1, "first"), ("2024", 12, "twelfth")],
        kod_format="000",
    )
    result = parse_register_file(p)
    assert len(result.kodlistor) == 1
    assert [r.kod for r in result.kodlistor[0].rows] == ["001", "012"]


@requires_openpyxl
def test_uppercase_xlsx_extension_picked_up_by_directory_parse(
    tmp_path: Path,
) -> None:
    p = tmp_path / "TEST.XLSX"
    _write_minimal_workbook(p)
    results = parse_directory(tmp_path)
    assert len(results) == 1
    assert results[0].source_file.name == "TEST.XLSX"


@requires_openpyxl
def test_directory_passed_as_file_rejected(tmp_path: Path) -> None:
    with pytest.raises(SosParseError, match="not a regular file"):
        parse_register_file(tmp_path)


@requires_openpyxl
def test_variable_sheet_without_variabelnamn_header_raises(tmp_path: Path) -> None:
    # An upstream rename or malformed delivery would leave the varsheet
    # without a Variabelnamn column. Silently returning zero variables
    # would hide the problem, so we fail fast here.
    p = tmp_path / "Test.xlsx"
    _write_minimal_workbook(p, var_header=["Foo", "Bar"])
    with pytest.raises(SosParseError, match="Variabelnamn"):
        parse_register_file(p)


@requires_openpyxl
def test_unsupported_xls_format_wrapped_as_sos_parse_error(tmp_path: Path) -> None:
    # openpyxl raises InvalidFileException for `.xls`/`.xlsb`; surface as
    # SosParseError so the parser's contract holds for common wrong inputs.
    p = tmp_path / "test.xls"
    p.write_bytes(b"")
    with pytest.raises(SosParseError, match="does not support"):
        parse_register_file(p)


# ---------------------------------------------------------------------------
# Integration tests against real deliveries
# ---------------------------------------------------------------------------


def _locate_sos_data() -> Path | None:
    """Find `regmeta/input_data/Socialstyrelsen/` — gitignored, so it may
    live in the main checkout rather than the current worktree.

    Honour `REGMETA_SOS_DATA` first. Otherwise walk upward so worktrees
    can find the sibling main checkout.
    """
    env = os.environ.get("REGMETA_SOS_DATA")
    if env:
        p = Path(env)
        return p if p.exists() else None

    anchor = Path(__file__).resolve()
    candidates = [anchor.parents[1] / "input_data" / "Socialstyrelsen"]
    for parent in anchor.parents:
        candidates.append(parent / "regmeta" / "input_data" / "Socialstyrelsen")

    for c in candidates:
        if c.exists() and list(c.glob("*.xlsx")):
            return c
    return None


SOS_DATA = _locate_sos_data()


requires_sos_data = pytest.mark.skipif(
    SOS_DATA is None or not HAS_OPENPYXL,
    reason=(
        "Socialstyrelsen input data not present (gitignored) or openpyxl "
        "missing; set REGMETA_SOS_DATA and install regmeta[xlsx] to run"
    ),
)


@requires_sos_data
def test_parses_every_register_without_error() -> None:
    results = parse_directory(SOS_DATA)
    assert len(results) >= 13, f"expected ≥13 registers, got {len(results)}"
    for r in results:
        assert isinstance(r, SosRegister)
        assert r.variables, f"{r.source_file.name}: no variables parsed"


@requires_sos_data
def test_skips_office_lock_files() -> None:
    results = parse_directory(SOS_DATA)
    for r in results:
        assert not r.source_file.name.startswith("~$")


@requires_sos_data
def test_par_shape() -> None:
    par = _find_register(parse_directory(SOS_DATA), "Patientregistret")
    assert par.dataset_name == "Patientregistret"
    assert par.dcat_ap.title_sv == "Patientregistret"
    assert par.dcat_ap.title_en == "National Patient Register"
    names = {d.name for d in par.deldatamangder}
    assert names == {"PAR_SV", "PAR_OV", "PAR_TV"}
    assert par.dcat_ap.legislation_sv
    assert "hälsodataregister" in par.dcat_ap.legislation_sv.lower()
    assert any(v.name == "HDIA" for v in par.variables)


@requires_sos_data
def test_mfr_kodlista_with_per_row_variable_column() -> None:
    mfr = _find_register(parse_directory(SOS_DATA), "Medicinska födelseregistret")
    butsatt = next(
        (k for k in mfr.kodlistor if k.variable_hint.lower() == "butsatt"), None
    )
    assert butsatt is not None, "Kodlista_butsatt missing"
    assert butsatt.rows
    assert butsatt.rows[0].variable_name == "BUTSATT"


@requires_sos_data
def test_unrecognised_kodlista_preserves_raw_rows() -> None:
    results = parse_directory(SOS_DATA)
    skipped = [k for r in results for k in r.kodlistor if not k.rows and k.raw_rows]
    assert skipped, "expected at least one kodlista with empty rows + raw_rows"


@requires_sos_data
def test_registers_without_deldatamangder_warn() -> None:
    results = parse_directory(SOS_DATA)
    missing = [r for r in results if not r.deldatamangder]
    assert missing, (
        "expected at least one register without deldatamängd sheet (LSS/BU/SOL)"
    )
    for r in missing:
        assert any("Deldatamängder" in w for w in r.warnings)


@requires_sos_data
def test_bu_phantom_rows_do_not_inflate_variable_count() -> None:
    bu = _find_register(parse_directory(SOS_DATA), "Insatser till barn och unga")
    assert 50 <= len(bu.variables) <= 2000, (
        f"BU variable count implausible: {len(bu.variables)}"
    )


@requires_sos_data
def test_dcat_ap_covers_expected_fields() -> None:
    par = _find_register(parse_directory(SOS_DATA), "Patientregistret")
    d: SosDcatAp = par.dcat_ap
    assert d.title_sv and d.title_en
    assert d.description_sv and d.description_en
    assert d.temporal_coverage_sv == "1964-"
    assert d.publisher_sv == "Socialstyrelsen"
    assert d.landing_page_sv and d.landing_page_sv.startswith("https://")
    assert d.access_rights_sv


@requires_sos_data
def test_variable_fields_populated() -> None:
    par = _find_register(parse_directory(SOS_DATA), "Patientregistret")
    hdia = next(v for v in par.variables if v.name == "HDIA")
    assert hdia.deldatamangd in {"PAR_SV", "PAR_OV", "PAR_TV"}
    assert hdia.label
    assert hdia.data_type
    external_refs = [
        v.external_classification for v in par.variables if v.external_classification
    ]
    assert any("icd" in r.lower() for r in external_refs)


# ---------------------------------------------------------------------------
# helpers
# ---------------------------------------------------------------------------


def _find_register(results: list[SosRegister], dataset_name: str) -> SosRegister:
    for r in results:
        if r.dataset_name and dataset_name.lower() in r.dataset_name.lower():
            return r
    raise AssertionError(f"register {dataset_name!r} not in parse results")
