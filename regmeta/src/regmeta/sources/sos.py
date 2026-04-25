"""Parser for Socialstyrelsen metadata Excel deliveries.

Each Socialstyrelsen register is published as a standalone .xlsx workbook
with a consistent-but-not-uniform set of sheets. This module reads one
workbook and returns a `SosRegister` — a structured, DB-schema-independent
representation suitable for downstream DB ingestion or docs generation.

Known shape (derived from the 13 registers currently distributed):

    Generell information        — template & dataset version, contact
    Metadata-Datamängd (DCAT-AP) — register-level DCAT-AP metadata
    Deldatamängder och datavyer — subset/view descriptions (optional)
    Metadata - Variabelnivå     — variable rows (16 standard columns)
    Kodlista_*                  — per-variable value sets (optional)
    Kvalitet_*                  — free-form quality notes (LMED only)

Sheet names vary in case, whitespace, and punctuation; `_find_sheet`
matches on normalised tokens. Workbook files beginning with `~$` are
Microsoft Office lock files and are rejected up front.
"""

from __future__ import annotations

import re
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

# ---------------------------------------------------------------------------
# Output types
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class SosDcatAp:
    """Register-level DCAT-AP metadata. All fields optional — older
    template versions or partial deliveries may omit any of them.

    `extras` holds rows whose Swedish attribute name we don't map to a
    known field, so the full sheet content survives parsing.
    """

    title_sv: str | None = None
    title_en: str | None = None
    description_sv: str | None = None
    description_en: str | None = None
    temporal_coverage_sv: str | None = None
    temporal_coverage_en: str | None = None
    geographic_coverage_sv: str | None = None
    geographic_coverage_en: str | None = None
    population_sv: str | None = None
    population_en: str | None = None
    update_frequency_sv: str | None = None
    update_frequency_en: str | None = None
    publisher_sv: str | None = None
    publisher_en: str | None = None
    contact_sv: str | None = None
    contact_en: str | None = None
    documentation_url_sv: str | None = None
    documentation_url_en: str | None = None
    landing_page_sv: str | None = None
    landing_page_en: str | None = None
    access_url_sv: str | None = None
    access_url_en: str | None = None
    access_rights_sv: str | None = None
    access_rights_en: str | None = None
    legislation_sv: str | None = None
    legislation_en: str | None = None
    extras: dict[str, str] = field(default_factory=dict)


@dataclass(frozen=True)
class SosDeldatamangd:
    """One subset/view within a register. From `Deldatamängder…` sheet.

    Some registers (LSS, BU, SOL) lack this sheet entirely; the caller is
    expected to synthesise a single implicit deldatamängd for those.
    """

    name: str
    label: str | None
    description: str | None
    data_from: int | None
    data_to: int | None
    update_frequency: str | None
    aggregation_level: str | None


@dataclass(frozen=True)
class SosVariable:
    """One variable occurrence in a register (row in Metadata - Variabelnivå).

    Identity is `(deldatamangd, name)`. The same variable name can appear
    under multiple deldatamängder within the same register, and across
    registers — uniqueness is not guaranteed even within a single file.
    """

    deldatamangd: str | None
    name: str
    label: str | None
    description: str | None
    object_type: str | None
    value_set_text: str | None  # raw `Värdemängd` free-text
    external_classification: str | None  # raw `Länk kodverk`
    data_type: str | None
    is_join_variable: str | None
    join_description: str | None
    presentation_order: int | None
    data_from: int | None
    data_to: int | None
    quality_note: str | None
    origin: str | None
    source_detail: str | None


@dataclass(frozen=True)
class SosKodlistaRow:
    tidsperiod: str | None
    kod: str
    beskrivning: str | None
    variable_name: str | None = (
        None  # set only when sheet has a per-row Variabelnamn column
    )


@dataclass(frozen=True)
class SosKodlista:
    """Value set from a `Kodlista_*` sheet. Mapping to a variable is by
    sheet-name suffix (e.g. `Kodlista_DIAGNOS` → variable `DIAGNOS`).
    The caller is responsible for resolution — not guaranteed 1:1.

    `rows` holds structured (Tidsperiod, Kod, Beskrivning) entries. Sheets
    that don't match the standard header shape (recoding tables, hospital
    directories, ICD mapping tables etc.) parse with empty `rows` — the
    raw content is preserved in `raw_rows` for downstream custom handling.
    """

    sheet_name: str
    variable_hint: str  # suffix after `Kodlista_`
    codeset_name: str | None  # from "Kodverk" row, if present
    variable_header: str | None  # from "Variabelnamn" row, if present
    background: str | None  # from "Bakgrund" row, if present
    rows: tuple[SosKodlistaRow, ...]
    raw_rows: tuple[tuple[Any, ...], ...] = ()


@dataclass(frozen=True)
class SosQualitySheet:
    """A `Kvalitet_*` sheet captured verbatim. LMED uses these for
    register-level quality narrative. Rows are kept as raw tuples; no
    further structure is assumed."""

    sheet_name: str
    rows: tuple[tuple[Any, ...], ...]


@dataclass(frozen=True)
class SosRegister:
    source_file: Path
    dataset_name: str | None
    dataset_version: str | None
    dataset_date: date | None
    template_version: str | None
    template_date: date | None
    contact_email: str | None
    dcat_ap: SosDcatAp
    deldatamangder: tuple[SosDeldatamangd, ...]
    variables: tuple[SosVariable, ...]
    kodlistor: tuple[SosKodlista, ...]
    quality_sheets: tuple[SosQualitySheet, ...]
    warnings: tuple[str, ...]


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------


class SosParseError(Exception):
    """Raised when the workbook cannot be read or is missing required sheets."""


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


def parse_register_file(path: Path | str) -> SosRegister:
    """Read one Socialstyrelsen register workbook and return structured
    metadata. Raises `SosParseError` on unreadable / unrecognised files."""

    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover - trivial
        raise SosParseError(
            "openpyxl is required for Socialstyrelsen parsing; "
            "install with `pip install regmeta[xlsx]`"
        ) from exc

    p = Path(path)
    if p.name.startswith("~$"):
        raise SosParseError(f"{p.name} is an Office lock file; skip")
    if not p.is_file():
        raise SosParseError(f"{p} is not a regular file (missing or a directory)")
    try:
        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
    except zipfile.BadZipFile as exc:
        raise SosParseError(f"{p.name} is not a valid .xlsx file") from exc
    except openpyxl.utils.exceptions.InvalidFileException as exc:
        # `.xls`, `.xlsb`, and other formats openpyxl doesn't support.
        raise SosParseError(
            f"{p.name}: openpyxl does not support this file format "
            "(only .xlsx/.xlsm/.xltx/.xltm)"
        ) from exc
    except (OSError, ValueError, KeyError) as exc:
        # openpyxl can raise these on partially corrupt files (truncated XML,
        # missing relationships, unexpected schema). Wrap so callers see a
        # uniform error type.
        raise SosParseError(
            f"{p.name} could not be read as a valid .xlsx file: {exc}"
        ) from exc

    try:
        warnings: list[str] = []
        norm_sheets = {_normalise(n): n for n in wb.sheetnames}

        generell = _find_sheet(norm_sheets, ["generell", "information"])
        dcat = _find_sheet(norm_sheets, ["datamängd", "dcat"]) or _find_sheet(
            norm_sheets, ["metadata", "datamängd"]
        )
        deldat = (
            _find_sheet(norm_sheets, ["deldatamängder", "datavyer"])
            or _find_sheet(norm_sheets, ["metadata", "deldatamängder"])
            or _find_sheet(norm_sheets, ["deldatamängder"])
        )
        varsheet = _find_sheet(
            norm_sheets, ["metadata", "variabelnivå"]
        ) or _find_sheet(norm_sheets, ["metadata", "variabler"])

        if varsheet is None:
            raise SosParseError(f"{p.name}: no variable-level sheet found")

        gen = _parse_generell(wb[generell]) if generell else {}
        dcat_ap = _parse_dcat_ap(wb[dcat]) if dcat else SosDcatAp()
        deldatamangder = tuple(_parse_deldatamangder(wb[deldat])) if deldat else ()
        variables = tuple(_parse_variables(wb[varsheet]))

        kodlistor: list[SosKodlista] = []
        quality_sheets: list[SosQualitySheet] = []
        for sheet_name in wb.sheetnames:
            low = sheet_name.lower()
            if low.startswith("kodlista"):
                try:
                    kod, kod_warnings = _parse_kodlista(wb[sheet_name])
                    kodlistor.append(kod)
                    warnings.extend(kod_warnings)
                except Exception as exc:
                    warnings.append(f"kodlista {sheet_name!r}: {exc}")
            elif low.startswith("kvalitet"):
                quality_sheets.append(_parse_quality_sheet(wb[sheet_name]))

        if generell is None:
            warnings.append("missing Generell information sheet")
        if dcat is None:
            warnings.append("missing DCAT-AP sheet")
        if deldat is None:
            warnings.append("missing Deldatamängder sheet (implicit single subset)")

        return SosRegister(
            source_file=p,
            dataset_name=gen.get("dataset_name"),
            dataset_version=gen.get("dataset_version"),
            dataset_date=gen.get("dataset_date"),
            template_version=gen.get("template_version"),
            template_date=gen.get("template_date"),
            contact_email=gen.get("contact_email"),
            dcat_ap=dcat_ap,
            deldatamangder=deldatamangder,
            variables=variables,
            kodlistor=tuple(kodlistor),
            quality_sheets=tuple(quality_sheets),
            warnings=tuple(warnings),
        )
    finally:
        wb.close()


def parse_directory(directory: Path | str) -> list[SosRegister]:
    """Parse every `.xlsx` file in a directory, skipping Office lock files.
    Halts on the first parse failure (raises `SosParseError`); call per file
    if you need to collect errors instead."""

    d = Path(directory)
    out: list[SosRegister] = []
    for f in sorted(d.iterdir()):
        if not f.is_file():
            continue
        # Case-insensitive: some deliveries arrive as `.XLSX` on case-sensitive
        # filesystems, and a strict `*.xlsx` glob would skip them silently.
        if f.suffix.lower() != ".xlsx":
            continue
        if f.name.startswith("~$"):
            continue
        out.append(parse_register_file(f))
    return out


# ---------------------------------------------------------------------------
# Sheet helpers
# ---------------------------------------------------------------------------


def _normalise(s: str) -> str:
    return re.sub(r"[\s_\-()]+", "", s).lower()


def _find_sheet(norm_sheets: dict[str, str], tokens: list[str]) -> str | None:
    """Return the first original sheet name whose normalised form contains
    every token in `tokens` (also normalised). Caller is expected to build
    `norm_sheets` once via `{_normalise(n): n for n in wb.sheetnames}` so
    repeated lookups don't re-normalise."""
    wanted = [_normalise(t) for t in tokens]
    for norm, original in norm_sheets.items():
        if all(t in norm for t in wanted):
            return original
    return None


def _row_iter(ws: Any, start: int = 1) -> Iterable[tuple[Any, ...]]:
    """Yield rows starting at `start`, stopping after a long empty tail.
    openpyxl's `max_row` is unreliable (phantom rows in some deliveries)."""
    empty_streak = 0
    empty_limit = 50
    for row in ws.iter_rows(min_row=start, values_only=True):
        if any(v is not None and str(v).strip() for v in row):
            empty_streak = 0
            yield row
        else:
            empty_streak += 1
            if empty_streak >= empty_limit:
                break


def _cell_row_iter(ws: Any, start: int = 1) -> Iterable[tuple[Any, ...]]:
    """Like `_row_iter` but yields tuples of openpyxl cell objects, so
    callers can inspect formatting (e.g. number_format on code columns)."""
    empty_streak = 0
    empty_limit = 50
    for cells in ws.iter_rows(min_row=start, values_only=False):
        if any(c.value is not None and str(c.value).strip() for c in cells):
            empty_streak = 0
            yield tuple(cells)
        else:
            empty_streak += 1
            if empty_streak >= empty_limit:
                break


def _at(row: tuple[Any, ...], idx: int | None) -> Any:
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _pick(row: tuple[Any, ...], col_map: dict[str, int], field_name: str) -> Any:
    """Look up `field_name` in `col_map` and return the row value at that
    index, or None if the column is absent or short. Convenience for
    header-mapped sheet parsers."""
    return _at(row, col_map.get(field_name))


_PURE_ZERO_FMT = re.compile(r"^0+$")


def _format_code(cell: Any) -> str | None:
    """Render a code-column cell to a string, preserving leading zeros from
    Excel display formatting. Excel may store '001' as the integer 1 with
    number_format '000'; without consulting the format we'd silently emit
    '1' and corrupt code identity for downstream joins."""
    v = cell.value
    if v is None:
        return None
    if isinstance(v, bool):  # bool is a subclass of int — handle first
        return str(v)
    if isinstance(v, (int, float)):
        if isinstance(v, float):
            if not v.is_integer():
                return str(v)
            v = int(v)
        fmt = cell.number_format or ""
        if _PURE_ZERO_FMT.fullmatch(fmt):
            return str(v).zfill(len(fmt))
        return str(v)
    s = str(v).strip()
    return s or None


def _clean(v: Any) -> str | None:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s or None
    return str(v)


def _as_int(v: Any) -> int | None:
    if v is None or v == "":
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return int(v) if v.is_integer() else None
    try:
        return int(str(v).strip())
    except (TypeError, ValueError):
        return None


def _as_date(v: Any) -> date | None:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


# ---------------------------------------------------------------------------
# Sheet-specific parsers
# ---------------------------------------------------------------------------


def _parse_generell(ws: Any) -> dict[str, Any]:
    """Scan `Generell information`. Layout is key–value pairs scattered
    across roughly 25 rows with section headings; we key on the label in
    column B (index 1) and read the value from column C (index 2)."""

    out: dict[str, Any] = {}
    rows = list(ws.iter_rows(values_only=True))
    section: str | None = None
    for row in rows:
        if len(row) < 2:
            continue
        label = _clean(row[1]) if len(row) > 1 else None
        value = _clean(row[2]) if len(row) > 2 else None
        raw_value = row[2] if len(row) > 2 else None

        if label and value is None:
            low = label.lower()
            # Some deliveries have "metadatat" (typo) instead of "metadata";
            # match on the distinguishing tail words instead of the exact phrase.
            if "metadatamallen" in low:
                section = "template"
            elif "datamängden" in low and "version" in low:
                section = "dataset"
            continue

        if not label or value is None:
            continue

        low = label.lower()
        if section == "template" and low.startswith("version"):
            out["template_version"] = value
        elif section == "template" and low.startswith("datum"):
            out["template_date"] = _as_date(raw_value)
        elif section == "dataset" and low.startswith("datamängd"):
            out["dataset_name"] = value
        elif section == "dataset" and low.startswith("version"):
            out["dataset_version"] = value
        elif section == "dataset" and low.startswith("datum"):
            out["dataset_date"] = _as_date(raw_value)
        elif "e-post" in low or low == "e-post:":
            out["contact_email"] = value
    return out


# DCAT-AP attribute (Swedish) → internal field stem. We store both SV and
# EN columns as separate `<stem>_sv` / `<stem>_en` fields.
_DCAT_MAP = {
    "titel": "title",
    "beskrivning": "description",
    "tidsperiod": "temporal_coverage",
    "namngivet geografiskt område": "geographic_coverage",
    "population": "population",
    "uppdateringsfrekvens": "update_frequency",
    "utgivare": "publisher",
    "kontaktuppgift": "contact",
    "dokumentation": "documentation_url",
    "ingångssida": "landing_page",
    "webbadress för åtkomst": "access_url",
    "åtkomsträttigheter": "access_rights",
    "tillämplig lagstiftning": "legislation",
}


def _parse_dcat_ap(ws: Any) -> SosDcatAp:
    fields: dict[str, str | None] = {}
    extras: dict[str, str] = {}
    first = True
    for row in _row_iter(ws):
        if first:
            first = False
            # First row is column headers (Attribut SoS-metadata | ... | Svenska | Engelska)
            continue
        if len(row) < 3:
            continue
        attr = _clean(row[0])
        sv = _clean(row[2]) if len(row) > 2 else None
        en = _clean(row[3]) if len(row) > 3 else None
        if not attr:
            continue
        stem = _DCAT_MAP.get(attr.lower())
        if stem is None:
            # Capture unrecognised rows for inspection; value preference SV > EN
            value = sv or en or ""
            if value:
                extras[attr] = value
            continue
        if sv is not None:
            fields[f"{stem}_sv"] = sv
        if en is not None:
            fields[f"{stem}_en"] = en
    return SosDcatAp(**fields, extras=extras)


_DELDATAMANGD_HEADERS = {
    "deldatamängdsnamn": "name",
    "deldatamängdsetikett": "label",
    "deldatamängbeskrivning": "description",
    "deldatamängdsbeskrivning": "description",
    "data från": "data_from",
    "data till": "data_to",
    "uppdateringsfrekvens": "update_frequency",
    "aggregeringsnivå": "aggregation_level",
}


def _parse_deldatamangder(ws: Any) -> Iterable[SosDeldatamangd]:
    rows = _row_iter(ws)
    header = next(rows, None)
    if header is None:
        return
    col_map: dict[str, int] = {}
    for i, h in enumerate(header):
        stem = _DELDATAMANGD_HEADERS.get(_clean(h or "").lower() if h else "")
        if stem:
            col_map[stem] = i

    if "name" not in col_map:
        return  # not a deldatamängd sheet shape; silently skip

    for row in rows:
        name = _clean(_pick(row, col_map, "name"))
        if not name:
            continue
        yield SosDeldatamangd(
            name=name,
            label=_clean(_pick(row, col_map, "label")),
            description=_clean(_pick(row, col_map, "description")),
            data_from=_as_int(_pick(row, col_map, "data_from")),
            data_to=_as_int(_pick(row, col_map, "data_to")),
            update_frequency=_clean(_pick(row, col_map, "update_frequency")),
            aggregation_level=_clean(_pick(row, col_map, "aggregation_level")),
        )


_VAR_HEADERS = {
    "deldatamängdsnamn": "deldatamangd",
    # BU splits deldatamängd into dataset + view; we keep the view name
    # ("Datavynamn") as the deldatamängd identity and drop the parent
    # ("Datamängdsnamn") since it duplicates the register-level name.
    "datavynamn": "deldatamangd",
    "variabelnamn": "name",
    "variabeletikett": "label",
    "variabelbeskrivning": "description",
    "objekttyp": "object_type",
    "värdemängd": "value_set_text",
    "länk kodverk": "external_classification",
    "datatyp": "data_type",
    "kopplingsvariabel": "is_join_variable",
    "kopplingsbeskrivning": "join_description",
    "presentationsordning": "presentation_order",
    "data från": "data_from",
    "data till": "data_to",
    "kvalitetsanmärkning": "quality_note",
    "ursprung": "origin",
    "specificera källa": "source_detail",
}


def _parse_variables(ws: Any) -> Iterable[SosVariable]:
    rows = _row_iter(ws)
    header = next(rows, None)
    if header is None:
        raise SosParseError(
            f"variable sheet {ws.title!r} is empty; cannot extract variables"
        )
    col_map: dict[str, int] = {}
    for i, h in enumerate(header):
        if not h:
            continue
        stem = _VAR_HEADERS.get(_clean(h).lower())
        if stem:
            col_map[stem] = i

    if "name" not in col_map:
        # Without a Variabelnamn column we silently return zero rows, hiding
        # an upstream rename or malformed delivery. Fail loudly instead.
        header_cols = ", ".join(repr(h) for h in header if h) or "(none)"
        raise SosParseError(
            f"variable sheet {ws.title!r} is missing a 'Variabelnamn' header; "
            f"found columns: {header_cols}"
        )

    for row in rows:
        name = _clean(_pick(row, col_map, "name"))
        if not name:
            continue
        yield SosVariable(
            deldatamangd=_clean(_pick(row, col_map, "deldatamangd")),
            name=name,
            label=_clean(_pick(row, col_map, "label")),
            description=_clean(_pick(row, col_map, "description")),
            object_type=_clean(_pick(row, col_map, "object_type")),
            value_set_text=_clean(_pick(row, col_map, "value_set_text")),
            external_classification=_clean(
                _pick(row, col_map, "external_classification")
            ),
            data_type=_clean(_pick(row, col_map, "data_type")),
            is_join_variable=_clean(_pick(row, col_map, "is_join_variable")),
            join_description=_clean(_pick(row, col_map, "join_description")),
            presentation_order=_as_int(_pick(row, col_map, "presentation_order")),
            data_from=_as_int(_pick(row, col_map, "data_from")),
            data_to=_as_int(_pick(row, col_map, "data_to")),
            quality_note=_clean(_pick(row, col_map, "quality_note")),
            origin=_clean(_pick(row, col_map, "origin")),
            source_detail=_clean(_pick(row, col_map, "source_detail")),
        )


def _parse_kodlista(ws: Any) -> tuple[SosKodlista, list[str]]:
    """A Kodlista sheet has a preamble (rows labelled Kodverk / Variabelnamn
    / Bakgrund) then a header row with (Tidsperiod, Kod, Beskrivning) and
    data rows beneath. Some sheets omit the preamble."""

    codeset_name: str | None = None
    variable_header: str | None = None
    background: str | None = None
    data_rows: list[SosKodlistaRow] = []
    raw_rows: list[tuple[Any, ...]] = []
    warnings: list[str] = []

    sheet_name = ws.title
    suffix = sheet_name.split("_", 1)[1] if "_" in sheet_name else sheet_name
    suffix = suffix.split("!", 1)[0].strip()

    col_tp: int | None = None
    col_kod: int | None = None
    col_desc: int | None = None
    col_var: int | None = None
    # forward-fill: some sheets put the period once on a header row above
    # rows that leave it blank
    last_tidsperiod: str | None = None
    # iterate cells (not just values) so the kod column can preserve leading
    # zeros from number_format
    all_cell_rows = list(_cell_row_iter(ws))
    for cells in all_cell_rows:
        row = tuple(c.value for c in cells)
        first = _clean(row[0]) if row else None
        second = _clean(row[1]) if len(row) > 1 else None

        if col_tp is None:
            # Detect the column-header row by looking for "Tidsperiod" + "Kod"
            # anywhere in the row. MFR-style sheets also carry a Variabelnamn
            # column (per-row variable) in addition to the Kodverk preamble
            # shape used by PAR et al.
            positions: dict[str, int] = {}
            for i, h in enumerate(row):
                hl = (_clean(h) or "").lower()
                if hl.startswith("tidsperiod"):
                    positions["tp"] = i
                elif hl == "kod":
                    positions["kod"] = i
                elif hl.startswith("beskrivning") or hl.startswith("betydelse"):
                    positions["desc"] = i
                elif hl == "variabelnamn":
                    positions["var"] = i
            if "tp" in positions and "kod" in positions:
                col_tp = positions["tp"]
                col_kod = positions["kod"]
                col_desc = positions.get("desc")
                col_var = positions.get("var")
                continue

            # Preamble rows (PAR-style): "Kodverk", "Variabelnamn", "Bakgrund"
            # appear in col 0 with the value in col 1.
            if first:
                key = first.lower()
                if key == "kodverk":
                    codeset_name = second
                    continue
                if key == "variabelnamn":
                    variable_header = second
                    continue
                if key == "bakgrund":
                    background = second
                    continue
            continue

        tp_val = _clean(_at(row, col_tp))
        kod_str = (
            _format_code(cells[col_kod])
            if col_kod is not None and col_kod < len(cells)
            else None
        )

        # Rows carrying only a Tidsperiod (no code) act as a section header
        # for the rows beneath them; remember and forward-fill.
        if tp_val and not kod_str:
            last_tidsperiod = tp_val
            continue
        if not kod_str:
            continue

        data_rows.append(
            SosKodlistaRow(
                tidsperiod=tp_val or last_tidsperiod,
                kod=kod_str,
                beskrivning=_clean(_at(row, col_desc)),
                variable_name=(
                    _clean(_at(row, col_var)) if col_var is not None else None
                ),
            )
        )

    if col_tp is None:
        warnings.append(
            f"kodlista {sheet_name!r}: no Tidsperiod/Kod header row found; "
            "structured rows skipped (raw content preserved)"
        )
        raw_rows = [tuple(c.value for c in cells) for cells in all_cell_rows]

    return (
        SosKodlista(
            sheet_name=sheet_name,
            variable_hint=suffix,
            codeset_name=codeset_name,
            variable_header=variable_header,
            background=background,
            rows=tuple(data_rows),
            raw_rows=tuple(raw_rows),
        ),
        warnings,
    )


def _parse_quality_sheet(ws: Any) -> SosQualitySheet:
    rows: list[tuple[Any, ...]] = []
    for row in _row_iter(ws):
        rows.append(tuple(row))
    return SosQualitySheet(sheet_name=ws.title, rows=tuple(rows))
