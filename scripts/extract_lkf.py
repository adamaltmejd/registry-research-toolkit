#!/usr/bin/env python3
"""Extract per-year LKF (län/kommun/församling) canonical codes from SCB.

SCB publishes one LKF snapshot per year as either lkf{YEAR}.xls (older) or
lkf{YEAR}.xlsx (newer) under
``contentassets/13ec5841d80045498d960d456e87ea78/``. Each file is a flat
table at row 12+:

    Län | Länsnamn | Kommun | Kommunnamn | Församling | Församlingsnamn

Each data row contributes three LKF codes — the 2-digit län, the 4-digit
kommun, and the 6-digit församling — all of which appear together in
register data. The output is one canonical CSV per year, ready to wire
into ``classifications.toml`` as ``valid_codes_file = "lkf{YEAR}.csv"``.

Coverage (as of probing in 2026-04):

    1974–2015  knkodnyckel.xls — kommun (4-digit) + derived län (2-digit).
               Parishes (6-digit) are not in this file; they remain a gap.
               12 period-snapshots cover this range; län names are year-aware
               to handle the 1997 (Skåne, Dalarna) and 1998 (Västra Götaland)
               reforms.
    2016–2026  per-year lkf{year}.xls/.xlsx — full LKF (län + kommun + parish).
               2020 and 2026 have non-standard filenames (URL_OVERRIDES).
    2015 PDF   downloadable via --download-pdfs but not parsed here. OCR it
               separately if you want parishes for 2015.
    Pre-1974   not addressable from SCB's modern downloads. REGINA (the SCB
               regional-indelningar tracker) covers 1952+ but isn't a file.

Run:
    uv run --with openpyxl --with xlrd python scripts/extract_lkf.py \\
        --out regmeta/input_data/classifications/

Add --download-pdfs to also fetch the PDF editions for 2015–2021.
Add --emit-toml to print starter [[classification]] entries on stdout.

Usage:
    uv run --with openpyxl --with xlrd python scripts/extract_lkf.py \\
        --out regmeta/input_data/classifications/

Add --emit-toml to print seed entries for classifications.toml on stderr.
"""

from __future__ import annotations

import argparse
import csv
import sys
import urllib.request
from pathlib import Path

# Years we know are downloadable as full LKF XLS/XLSX (län + kommun + parish).
YEARS_AVAILABLE = list(range(2016, 2027))

# Years served from the kommun-history file (kommun + derived län only — no
# parishes). Together with YEARS_AVAILABLE this gives 1974–2026 coverage.
KN_YEARS = list(range(1974, 2016))

BASE_URL = "https://www.scb.se/contentassets/13ec5841d80045498d960d456e87ea78"
URL_OVERRIDES = {
    # 2020 was published as a "justerad" (corrected) edition with a non-standard
    # filename; the standard /lkf2020.xls path returns 404.
    2020: f"{BASE_URL}/lkf2020_justerad-1.xls",
    # 2026 sits in a date-stamped subdirectory.
    2026: f"{BASE_URL}/2025-06-19/lkf2026.xlsx",
}

# Kommun-history file: 12 period-snapshots from 1974 onwards in column pairs.
KNKODNYCKEL_URL = (
    "https://www.scb.se/contentassets/6a74a52b28994e2bbe23dcdd6754987c/knkodnyckel.xls"
)
KNKODNYCKEL_PERIODS: list[tuple[int, int, int]] = [
    # (year_from, year_to_inclusive, kod_col)
    (1974, 1976, 0),
    (1977, 1979, 4),
    (1980, 1982, 8),
    (1983, 1991, 12),
    (1992, 1994, 16),
    (1995, 1996, 20),
    (1997, 1997, 24),
    (1998, 1998, 28),
    (1999, 2002, 32),
    (2003, 2006, 36),
    (2007, 2007, 40),
    (2008, 2099, 43),
]

# Län names. The post-1998 set is stable; pre-reform names are hardcoded.
# Reforms: 1997 (Skåne, Dalarnas), 1998 (Västra Götaland).
_LAN_BASE = {
    "01": "Stockholms län",
    "03": "Uppsala län",
    "04": "Södermanlands län",
    "05": "Östergötlands län",
    "06": "Jönköpings län",
    "07": "Kronobergs län",
    "08": "Kalmar län",
    "09": "Gotlands län",
    "10": "Blekinge län",
    "13": "Hallands län",
    "17": "Värmlands län",
    "18": "Örebro län",
    "19": "Västmanlands län",
    "21": "Gävleborgs län",
    "22": "Västernorrlands län",
    "23": "Jämtlands län",
    "24": "Västerbottens län",
    "25": "Norrbottens län",
}


def lan_names_for(year: int) -> dict[str, str]:
    """Return the län-code → name map valid for the given year."""
    if year < 1997:
        return {
            **_LAN_BASE,
            "11": "Kristianstads län",
            "12": "Malmöhus län",
            "14": "Göteborgs och Bohus län",
            "15": "Älvsborgs län",
            "16": "Skaraborgs län",
            "20": "Kopparbergs län",
        }
    if year == 1997:  # Kristianstad+Malmöhus → Skåne; Kopparberg → Dalarna
        return {
            **_LAN_BASE,
            "12": "Skåne län",
            "14": "Göteborgs och Bohus län",
            "15": "Älvsborgs län",
            "16": "Skaraborgs län",
            "20": "Dalarnas län",
        }
    # 1998+: GoB + Älvsborg + parts of Skaraborg → Västra Götaland
    return {
        **_LAN_BASE,
        "12": "Skåne län",
        "14": "Västra Götalands län",
        "20": "Dalarnas län",
    }


# PDFs available at the standard URL pattern. 2015 is PDF-only; 2016–2021 also
# have PDFs (kept for cross-checking against XLS extraction).
PDF_YEARS = list(range(2015, 2022))
PDF_OVERRIDES = {
    # 2020 PDF is also _justerad.
    2020: f"{BASE_URL}/lkf2020_justerad.pdf",
}


def pdf_url_for(year: int) -> str:
    return PDF_OVERRIDES.get(year, f"{BASE_URL}/lkf{year}.pdf")


def url_for(year: int) -> str:
    if year in URL_OVERRIDES:
        return URL_OVERRIDES[year]
    ext = "xlsx" if year >= 2026 else "xls"
    return f"{BASE_URL}/lkf{year}.{ext}"


def download(year: int, dst: Path) -> Path:
    """Download lkf{year} into dst (skipped if already present)."""
    url = url_for(year)
    fname = url.rsplit("/", 1)[-1]
    out = dst / fname
    if out.exists():
        return out
    print(f"  downloading {url}", file=sys.stderr)
    urllib.request.urlretrieve(url, out)
    return out


def extract(year: int, src: Path) -> dict[str, str]:
    """Read one LKF snapshot file, return {vardekod: vardebenamning}.

    SCB has used at least three different layouts over the years (2018–19
    use a 2-column layout, 2021–22 a sparse 5-column layout, 2023+ a
    6-column table). To stay robust, the extractor doesn't try to anchor
    on any specific header — it walks every row and takes the first
    code-like cell (2/4/6-digit numeric string) paired with the next
    non-empty cell as the label. Works across all observed layouts.
    """
    if src.suffix == ".xlsx":
        import openpyxl

        wb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    else:
        import xlrd

        wb = xlrd.open_workbook(src)
        sh = wb.sheet_by_index(0)
        rows = [tuple(sh.row_values(i)) for i in range(sh.nrows)]

    out: dict[str, str] = {}
    for row in rows:
        i = 0
        while i < len(row):
            cell = row[i]
            if cell is None:
                i += 1
                continue
            code = str(cell).strip()
            if not (code.isdigit() and len(code) in (2, 4, 6)):
                i += 1
                continue
            # Found a code. Take the next non-empty cell as its label, then
            # advance past it. The 2023+ table layout has multiple codes per
            # row; vertical layouts have one but we still walk uniformly.
            label = ""
            label_idx = i + 1
            for j in range(i + 1, len(row)):
                if row[j] is None:
                    continue
                s = str(row[j]).strip()
                if s:
                    label = s
                    label_idx = j
                    break
            if label:
                out.setdefault(code, label)
            i = label_idx + 1
    return out


def extract_from_knkodnyckel(year: int, src: Path) -> dict[str, str]:
    """Return {kommun: name, lan: name} for a year served from knkodnyckel.xls.

    Knkodnyckel covers 1974+ as 12 period-snapshots in column pairs. It only
    has kommun codes — län codes are 2-digit prefixes of the kommun, named
    via lan_names_for() (year-aware to handle the 1997/1998 reforms).
    Församlings (6-digit) aren't here; pre-2016 parishes stay unmapped.
    """
    period_col = next(
        (col for lo, hi, col in KNKODNYCKEL_PERIODS if lo <= year <= hi),
        None,
    )
    if period_col is None:
        raise SystemExit(f"no knkodnyckel period covers year {year}")

    import xlrd

    wb = xlrd.open_workbook(src)
    sh = wb.sheet_by_index(0)

    out: dict[str, str] = {}
    seen_lan: set[str] = set()
    lan_lookup = lan_names_for(year)
    for i in range(2, sh.nrows):
        row = sh.row_values(i)
        if period_col + 1 >= len(row):
            continue
        kod_cell = row[period_col]
        namn_cell = row[period_col + 1]
        # xlrd returns numeric kommun codes as float — re-pad to 4 digits.
        if isinstance(kod_cell, float):
            kod = f"{int(kod_cell):04d}"
        else:
            kod = str(kod_cell).strip()
        namn = str(namn_cell).strip() if namn_cell else ""
        if not (kod.isdigit() and len(kod) == 4 and namn):
            continue
        # Names in knkodnyckel are uppercase; title-case them so they match
        # the modern XLS extraction style.
        out[kod] = namn.title().replace("Och ", "och ")
        seen_lan.add(kod[:2])

    for lan_code in sorted(seen_lan):
        out[lan_code] = lan_lookup.get(lan_code, f"Län {lan_code}")
    return out


def write_csv(year: int, codes: dict[str, str], out_dir: Path) -> Path:
    out = out_dir / f"lkf{year}.csv"
    with out.open("w", encoding="utf-8", newline="") as fh:
        w = csv.writer(fh)
        w.writerow(["vardekod", "vardebenamning"])
        for code, label in sorted(codes.items()):
            w.writerow([code, label])
    return out


def emit_toml_entry(year: int, prev_year: int | None) -> str:
    """Emit a [[classification]] block for LKF{year}."""
    # The vardemangdsversion strings are inconsistent across years — see
    # the existing LKF entry in classifications.toml for the variants.
    # The script prints a starter; reconcile against actual observed
    # strings before committing.
    sup = f'\nsupersedes = "LKF{prev_year}"' if prev_year else ""
    return f"""\
[[classification]]
short_name = "LKF{year}"
name = "Län, kommuner och församlingar {year}"
name_en = "Counties, municipalities and parishes {year}"
publisher = "SCB"
version = "{year}"
valid_from = {year}
valid_to = {year}
url = "https://www.scb.se/hitta-statistik/regional-statistik-och-kartor/regionala-indelningar/lan-och-kommuner/"{sup}
valid_codes_file = "lkf{year}.csv"
vardemangdsversion = [
  # TODO: add the year's vardemangdsversion variants here. Sample patterns:
  #   "LKF {year}-01-01/ Län, kommuner och församlingar"
  #   "LKF {year}-01-01/ Län, kommuner och församlingar "  (trailing space)
  #   "{year}-01-01/ Län, kommuner och församlingar"       (no LKF prefix, 2018+)
  #   "{year}-01-01 /Län, kommuner och församlingar"       (space before /, 2023+)
]
"""


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument(
        "--out",
        type=Path,
        default=Path("regmeta/input_data/classifications"),
        help="Where to write per-year CSVs.",
    )
    p.add_argument(
        "--cache",
        type=Path,
        default=Path("/tmp/scb_lkf"),
        help="Where to cache downloaded XLS/XLSX files.",
    )
    p.add_argument(
        "--years",
        type=int,
        nargs="*",
        default=KN_YEARS + YEARS_AVAILABLE,
        help=(
            "Specific years to extract (default: 1974–2026). 1974–2015 read "
            "kommun + derived län from knkodnyckel.xls (no parishes). "
            "2016–2026 read full LKF (län + kommun + parish) from per-year XLS."
        ),
    )
    p.add_argument(
        "--emit-toml",
        action="store_true",
        help="Print starter classifications.toml entries on stderr.",
    )
    p.add_argument(
        "--download-pdfs",
        action="store_true",
        help=(
            "Also download the PDF edition for each year listed in PDF_YEARS. "
            "Saved to the cache dir for later OCR/cross-checking; not parsed."
        ),
    )
    args = p.parse_args()

    args.out.mkdir(parents=True, exist_ok=True)
    args.cache.mkdir(parents=True, exist_ok=True)

    # knkodnyckel.xls is shared across all kn-years; download once.
    knkod_src: Path | None = None
    if any(y not in YEARS_AVAILABLE for y in args.years):
        knkod_src = args.cache / "knkodnyckel.xls"
        if not knkod_src.exists():
            print(f"  downloading {KNKODNYCKEL_URL}", file=sys.stderr)
            urllib.request.urlretrieve(KNKODNYCKEL_URL, knkod_src)

    prev_year: int | None = None
    for year in sorted(args.years):
        print(f"\n=== LKF {year} ===", file=sys.stderr)
        try:
            if year in YEARS_AVAILABLE:
                src = download(year, args.cache)
                codes = extract(year, src)
                source_note = src.name
            else:
                assert knkod_src is not None
                codes = extract_from_knkodnyckel(year, knkod_src)
                source_note = "knkodnyckel.xls (kommun + derived län only)"
            out = write_csv(year, codes, args.out)
            print(
                f"  wrote {out} ({len(codes)} codes from {source_note})",
                file=sys.stderr,
            )
            if args.emit_toml:
                sys.stdout.write(emit_toml_entry(year, prev_year))
                sys.stdout.write("\n")
            prev_year = year
        except Exception as exc:
            print(f"  FAILED: {exc}", file=sys.stderr)

    if args.download_pdfs:
        print("\n=== downloading PDFs ===", file=sys.stderr)
        for year in PDF_YEARS:
            url = pdf_url_for(year)
            dst = args.cache / f"lkf{year}.pdf"
            if dst.exists():
                print(f"  lkf{year}.pdf already cached", file=sys.stderr)
                continue
            try:
                urllib.request.urlretrieve(url, dst)
                print(
                    f"  lkf{year}.pdf: {dst.stat().st_size} bytes",
                    file=sys.stderr,
                )
            except Exception as exc:
                print(f"  lkf{year}.pdf FAILED: {exc}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
